#!/usr/bin/env /usr/bin/python3
"""
BIM Register Watchdog — monitors Submittal's and Design Files folders
across all 17 Samaya BIM projects for new/modified files and automatically
updates the relevant Excel registers.

Uses macOS FSEvents via ctypes (no third-party installs).
Uses Hermes notify script for Telegram alerts.

Author: Hermes AI
Schedule: Run via cron every 2 minutes (bim_watchdog --daemon for continuous)
"""
import os
import sys
import json
import time
import ctypes
import ctypes.util
import hashlib
import logging
import subprocess
import openpyxl
from datetime import datetime
from pathlib import Path

# ──────────────────────────────────────────────────────────────────────────────
# CONFIG
# ──────────────────────────────────────────────────────────────────────────────
BIM_UNIT = "/Users/mohamedessa/Library/CloudStorage/OneDrive-SAMAYAINVESTMENT/Samaya/Technical Office/Bim Unit"
REGISTERS_STATE = os.path.expanduser("~/.hermes/scripts/.watchdog_state.json")
HERMES_NOTIFY   = os.path.expanduser("~/.hermes/scripts/hermes_notify.sh")
LOG_FILE        = os.path.expanduser("~/.hermes/scripts/bim_watchdog.log")

SKIP_DIRS = {"Cache", "Archive", "archive", "_Archive", "Old", "old", "Backup", "backup", ".DS_Store"}
SKIP_EXTS = {".tmp", ".bak", ".lock", ".download"}

# Projects with a register folder
PROJECT_FOLDERS = {
    "Aseer-Museum":                     {"submittals": "Submittal's",          "design": "Design Files"},
    "El-Ghamama -(Qahwtna)":           {"submittals": "Submittal's",           "design": "Design Files"},
    "El-Ghamama Gift Shop(1)":         {"submittals": "Submittal's",           "design": "Design Files"},
    "El-Ghamama Gift Shop(2)":         {"submittals": "Submittal's",           "design": "Design Files"},
    "El-Ghamama Museum":                {"submittals": "Submittal's",            "design": "Design Files"},
    "El-Haramain Museum":               {"submittals": "Submittal's",           "design": "Design Files"},
    "Hera' Ghar":                      {"submittals": None,                    "design": "Design Files"},
    "Jabal Al-Noor Dispatch Building - Makkah": {"submittals": "Submittal's",  "design": "Design Files"},
    "Jabal Omar- Qahwtna (BassFl)":    {"submittals": "Submittal's",           "design": "Design Files"},
    "Jabal Omar- Retails 01 Hadaya Teiba الجلال و الجمال  (GND FL)": {"submittals": "Submittal's", "design": "Design Files"},
    "Jabal Omar- Retails 02 Hadaya Teiba معالم الحرمين (BASS FL)": {"submittals": "Submittal's", "design": "Design Files"},
    "Jabal Thawe --Thawr Heights":      {"submittals": "Submittal's",           "design": "Design Files"},
    "Khair El-Khalq Museum":           {"submittals": "Submittal's",           "design": "Design Files"},
    "Masjid Alnoor":                   {"submittals": None,                    "design": "Design Files"},
    "Prime Business Resort":           {"submittals": "Submittal's",           "design": "Design Files"},
    "Zamzam -Visitor Center":          {"submittals": "Submittal's",           "design": "Design Files"},
    "Zamzam Museum":                   {"submittals": "Submittal's",            "design": "Design Files"},
}

# Register types per project
REGISTER_TYPES = {
    "Aseer-Museum":       ["Drawing_Register", "Submittal_Register", "RFI_Register", "SI_Register", "NCR_Register", "Change_Order_Register", "Material_Register", "Invoice_Register", "Meeting_Minutes_Register", "Transmittal_Register", "Contract_Register", "Risk_Register", "Subcontractor_Register", "HSE_Register"],
    "Hera' Ghar":         ["Drawing_Register", "Submittal_Register"],
    "El-Ghamama Gift Shop(2)": ["Drawing_Register", "Submittal_Register", "RFI_Register", "SI_Register", "NCR_Register", "Change_Order_Register", "Material_Register", "Invoice_Register", "Meeting_Minutes_Register", "Transmittal_Register", "Contract_Register", "Risk_Register", "Subcontractor_Register", "HSE_Register"],
    "El-Ghamama Gift Shop(1)": ["Drawing_Register", "Submittal_Register"],
    "Zamzam Museum":      ["Drawing_Register", "Submittal_Register"],
    "Masjid Alnoor":     ["Drawing_Register"],
}

# Submittal disciplines
DISCIPLINES = {
    "Arch": "Architectural", "arch": "Architectural",
    "MEP": "MEP", "Mech": "Mechanical", "HVAC": "Mechanical",
    "Elec": "Electrical", "Plumb": "Plumbing", "Struct": "Structural",
    "AV": "AV / IT", "Furniture": "Furniture", "Furntiure": "Furniture",
    "RCP": "Architectural", "Flooring": "Architectural", "Partition": "Architectural",
}

# ──────────────────────────────────────────────────────────────────────────────
# LOGGING
# ──────────────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("bim_watchdog")

# ──────────────────────────────────────────────────────────────────────────────
# STATE
# ──────────────────────────────────────────────────────────────────────────────
def load_state():
    if os.path.exists(REGISTERS_STATE):
        try:
            with open(REGISTERS_STATE) as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def save_state(state):
    with open(REGISTERS_STATE, "w") as f:
        json.dump(state, f, indent=2, default=str)

def file_hash(path):
    try:
        with open(path, "rb") as f:
            return hashlib.md5(f.read(8192)).hexdigest()
    except Exception:
        return ""

# ──────────────────────────────────────────────────────────────────────────────
# FSEvents via ctypes
# ──────────────────────────────────────────────────────────────────────────────
class FSEventWatcher:
    """Lightweight FSEvents wrapper using macOS CoreServices ctypes."""

    def __init__(self, paths, callback, latency=1.0):
        self.paths = [os.path.abspath(p) for p in paths if os.path.exists(p)]
        self.callback = callback
        self.latency = latency
        self.running = False
        self._stream = None

        self._CFAbsoluteTimeGetCurrent = ctypes.CFUNCTYPE(ctypes.c_double)(
            ctypes.CFUNCTYPE(ctypes.c_double).from_address(
                ctypes.util.find_library("CoreFoundation")._handle  # noqa
            ) if False else None
        )

        # Load CoreServices
        self._cs = ctypes.CDLL(ctypes.util.find_library("CoreServices"))
        self._cs.CGEventSourceFlagSuppressClickFinderCount  # force load

    def _fsevents_callback(self, stream_ref, client_call_back_info, num_events, event_paths, event_flags, event_ids):
        for i in range(num_events):
            path = event_paths[i].decode("utf-8", errors="replace")
            flag = event_flags[i]
            # File events only (not dir)
            if flag & 0x10:  # kFSEventStreamEventFlagItemCreated
                self.callback(path, "created")
            if flag & 0x20:  # kFSEventStreamEventFlagItemModified
                self.callback(path, "modified")
            if flag & 0x40:  # kFSEventStreamEventFlagItemRemoved
                self.callback(path, "removed")
            if flag & 0x200:  # kFSEventStreamEventFlagItemRenamed
                self.callback(path, "renamed")
        return 0

    def start(self):
        if not self.paths:
            log.warning("No paths to watch")
            return

        callbacks_type = ctypes.CFUNCTYPE(
            None,
            ctypes.c_void_p, ctypes.c_void_p,
            ctypes.c_uint, ctypes.c_void_p,
            ctypes.c_uint, ctypes.c_uint
        )
        self._callbacks = callbacks_type(self._fsevents_callback)

        paths_arr = (ctypes.c_char_p * len(self.paths))()
        for i, p in enumerate(self.paths):
            paths_arr[i] = p.encode()

        flags = 0xFFF  # all events
        self._stream = self._cs.FSEventStreamCreate(
            None, self._callbacks, None,
            paths_arr, 1, self.latency
        )
        self._cs.FSEventStreamScheduleWithRunLoop(
            self._stream, ctypes.c_char_p(),  # current thread
            None
        )
        self._cs.FSEventStreamStart(self._stream)
        self.running = True
        log.info(f"FSEvents watching {len(self.paths)} paths")

    def stop(self):
        if self._stream:
            self._cs.FSEventStreamStop(self._stream)
            self._cs.FSEventStreamInvalidate(self._stream)
            self._cs.FSEventStreamRelease(self._stream)
        self.running = False

# ──────────────────────────────────────────────────────────────────────────────
# REGISTER UPDATE LOGIC
# ──────────────────────────────────────────────────────────────────────────────
def get_register_sheet_name(reg_name):
    """Map register filename to its data sheet name."""
    mapping = {
        "Drawing_Register": "Drawing", "Submittal_Register": "Submittal",
        "RFI_Register": "RFI", "SI_Register": "SI", "NCR_Register": "NCR",
        "Change_Order_Register": "ChangeOrder", "Material_Register": "Material",
        "Invoice_Register": "Invoice", "Meeting_Minutes_Register": "MeetingMinutes",
        "Transmittal_Register": "Transmittal", "Contract_Register": "Contract",
        "Risk_Register": "Risk", "Subcontractor_Register": "Subcontractor",
        "HSE_Register": "HSE",
    }
    return mapping.get(reg_name, reg_name.replace("_Register", ""))

def determine_discipline(path):
    """Guess discipline from folder name."""
    parts = Path(path).parts
    for p in parts:
        if p in DISCIPLINES:
            return DISCIPLINES[p]
    # Check submittal folders
    sub_dirs = {"HVAC": "Mechanical", "Electrical": "Electrical", "Plumbing": "Plumbing",
                "Structural": "Structural", "AV": "AV / IT", "IT": "IT"}
    for p in parts:
        if p in sub_dirs:
            return sub_dirs[p]
    return "General"

def next_drawing_number(reg_path, discipline="Architectural"):
    """Get next drawing number for a register."""
    try:
        wb = openpyxl.load_workbook(reg_path)
        ws = wb[get_register_sheet_name(os.path.basename(reg_path))]
        nums = []
        for row in ws.iter_rows(values_only=True):
            val = row[1] if len(row) > 1 else None  # Drawing # col
            if val and str(val).startswith(("GA-", "ST-", "DR-", "RCP-", "FL-", "TP-", "DF-", "AV-", "MEP-")):
                try:
                    nums.append(int(str(val).split("-")[-1]))
                except ValueError:
                    pass
        last = max(nums) if nums else 0
        return f"DR-{last+1:03d}"
    except Exception:
        return "DR-001"

def next_submittal_number(reg_path):
    """Get next submittal number."""
    try:
        wb = openpyxl.load_workbook(reg_path)
        ws = wb["Submittal"]
        nums = []
        for row in ws.iter_rows(values_only=True):
            val = row[1] if len(row) > 1 else None
            if val and str(val).startswith("SUB-"):
                try:
                    nums.append(int(str(val).split("-")[-1]))
                except ValueError:
                    pass
        last = max(nums) if nums else 0
        return f"SUB-{last+1:03d}"
    except Exception:
        return "SUB-001"

def is_valid_file(path):
    if not os.path.isfile(path):
        return False
    ext = os.path.splitext(path)[1].lower()
    if ext in SKIP_EXTS:
        return False
    if any(s in path for s in SKIP_DIRS):
        return False
    if os.path.basename(path).startswith("~"):
        return False
    return True

def is_valid_register(path):
    """Check if a register file is a valid xlsx (not a OneDrive cloud stub)."""
    if not os.path.exists(path):
        return False
    try:
        import zipfile
        with zipfile.ZipFile(path) as z:
            return len(z.filelist) > 0
    except Exception:
        return False

def add_to_drawing_register(reg_path, file_path, project_name):
    """Add a new file to the Drawing register."""
    if not os.path.exists(reg_path):
        return None
    if not is_valid_register(reg_path):
        log.warning(f"Drawing register is not accessible (OneDrive cloud stub?): {reg_path}")
        return None
    try:
        wb = openpyxl.load_workbook(reg_path)
        ws = wb["Drawing"]
        rel_path = file_path.split("Samaya/Technical Office/Bim Unit/")[-1] if "Samaya/Technical Office/Bim Unit/" in file_path else file_path
        discipline = determine_discipline(file_path)
        dwg_num = next_drawing_number(reg_path, discipline)
        date_str = datetime.now().strftime("%Y-%m-%d")
        new_row = [date_str, dwg_num, os.path.basename(file_path), discipline, "N/A", "N/A", "00", "New", rel_path, "Auto-added by watchdog"]
        ws.append(new_row)
        wb.save(reg_path)
        log.info(f"Drawing Register: added {dwg_num} for {os.path.basename(file_path)}")
        return dwg_num
    except Exception as e:
        log.error(f"Failed to update Drawing register: {e}")
        return None

def add_to_submittal_register(reg_path, file_path, project_name):
    """Add a new file to the Submittal register."""
    if not os.path.exists(reg_path):
        return None
    if not is_valid_register(reg_path):
        log.warning(f"Submittal register is not accessible (OneDrive cloud stub?): {reg_path}")
        return None
    try:
        wb = openpyxl.load_workbook(reg_path)
        ws = wb["Submittal"]
        rel_path = file_path.split("Samaya/Technical Office/Bim Unit/")[-1] if "Samaya/Technical Office/Bim Unit/" in file_path else file_path
        discipline = determine_discipline(file_path)
        sub_num = next_submittal_number(reg_path)
        date_str = datetime.now().strftime("%Y-%m-%d")
        new_row = [date_str, sub_num, os.path.basename(file_path), "Shop Drawing", "Samaya Investment", "Client", "Submitted", "-", rel_path, "Auto-added by watchdog"]
        ws.append(new_row)
        wb.save(reg_path)
        log.info(f"Submittal Register: added {sub_num} for {os.path.basename(file_path)}")
        return sub_num
    except Exception as e:
        log.error(f"Failed to update Submittal register: {e}")
        return None

def update_registers_for_project(project_name, changed_path, event_type):
    """Update all applicable registers for a project when a file changes."""
    project_dir = os.path.join(BIM_UNIT, project_name)
    regs_config = PROJECT_FOLDERS.get(project_name, {})
    results = []

    regs_to_update = REGISTER_TYPES.get(project_name, ["Drawing_Register", "Submittal_Register"])
    regs_base = os.path.join(project_dir, "Docs", "09_Registers")

    for reg_name in regs_to_update:
        reg_path = os.path.join(regs_base, f"{reg_name}.xlsx")
        if not os.path.exists(reg_path):
            continue

        if "Drawing" in reg_name:
            r = add_to_drawing_register(reg_path, changed_path, project_name)
            if r:
                results.append(f"{reg_name}: {r}")
        elif "Submittal" in reg_name:
            r = add_to_submittal_register(reg_path, changed_path, project_name)
            if r:
                results.append(f"{reg_name}: {r}")

    return results

# ──────────────────────────────────────────────────────────────────────────────
# EVENT HANDLER
# ──────────────────────────────────────────────────────────────────────────────
def on_file_event(path, event_type):
    if not is_valid_file(path):
        return

    # Find which project this belongs to
    for project_name in PROJECT_FOLDERS:
        if project_name in path:
            log.info(f"[{event_type.upper()}] {path}")
            results = update_registers_for_project(project_name, path, event_type)
            if results:
                notify_telegram(project_name, event_type, path, results)
            return

    log.debug(f"File event but no matching project: {path}")

def notify_telegram(project, event_type, path, results):
    """Send Telegram notification via Hermes notify script."""
    filename = os.path.basename(path)
    msg = (
        f"📁 *BIM Register Updated*\n"
        f"*Project:* {project}\n"
        f"*Event:* {event_type.upper()}\n"
        f"*File:* `{filename}`\n"
        f"*Registers updated:*\n"
    )
    for r in results:
        msg += f"  • `{r}`\n"
    msg += f"_Updated automatically by Hermes Watchdog_"

    try:
        subprocess.run(
            [HERMES_NOTIFY, "telegram", msg],
            capture_output=True, timeout=10
        )
    except Exception as e:
        log.warning(f"Telegram notify failed: {e}")

# ──────────────────────────────────────────────────────────────────────────────
# FSEvents polling (fallback for non-daemon mode)
# ──────────────────────────────────────────────────────────────────────────────
def get_all_watch_paths():
    """Build flat list of all directories to poll/watch."""
    paths = []
    for proj, cfg in PROJECT_FOLDERS.items():
        proj_dir = os.path.join(BIM_UNIT, proj)
        for key in ("submittals", "design"):
            folder = cfg.get(key)
            if folder:
                p = os.path.join(proj_dir, folder)
                if os.path.exists(p):
                    paths.append(p)
    return paths

def scan_and_update():
    """
    Full scan of all watch paths.
    Detects new/modified files vs last state and updates registers.
    Called by cron every 2 minutes.
    """
    state = load_state()
    events_found = []

    for proj, cfg in PROJECT_FOLDERS.items():
        proj_dir = os.path.join(BIM_UNIT, proj)
        for key in ("submittals", "design"):
            folder = cfg.get(key)
            if not folder:
                continue
            folder_path = os.path.join(proj_dir, folder)
            if not os.path.exists(folder_path):
                continue

            for dirpath, dirnames, filenames in os.walk(folder_path):
                dirnames[:] = [d for d in dirnames if d not in SKIP_DIRS]
                for fname in filenames:
                    fpath = os.path.join(dirpath, fname)
                    if not is_valid_file(fpath):
                        continue
                    key2 = fpath
                    fhash = file_hash(fpath)
                    prev = state.get(key2, {})
                    if not prev or prev.get("hash") != fhash:
                        state[key2] = {"hash": fhash, "mtime": os.path.getmtime(fpath)}
                        events_found.append((proj, fpath, "created/updated" if prev else "new"))

    save_state(state)

    if events_found:
        for proj, fpath, etype in events_found:
            log.info(f"[{etype.upper()}] {fpath}")
            results = update_registers_for_project(proj, fpath, etype)
            if results:
                notify_telegram(proj, etype, fpath, results)
        log.info(f"Scan complete: {len(events_found)} file(s) processed")
    else:
        log.info("Scan complete: no changes")

# ──────────────────────────────────────────────────────────────────────────────
# DAEMON MODE (FSEvents-based)
# ──────────────────────────────────────────────────────────────────────────────
def daemon_mode():
    paths = get_all_watch_paths()
    if not paths:
        log.error("No paths to watch")
        return

    log.info(f"Starting BIM watchdog daemon on {len(paths)} directories")
    watcher = FSEventWatcher(paths, on_file_event, latency=2.0)
    watcher.start()

    try:
        while True:
            time.sleep(2)
    except KeyboardInterrupt:
        log.info("Shutting down...")
        watcher.stop()

# ──────────────────────────────────────────────────────────────────────────────
# CLI
# ──────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--daemon":
        daemon_mode()
    elif len(sys.argv) > 1 and sys.argv[1] == "--scan":
        scan_and_update()
    else:
        print("Usage: bim_watchdog.py [--daemon|--scan]")
        print("  --daemon  Run continuously with FSEvents (macOS native)")
        print("  --scan    One-shot scan (for cron/polling mode)")
        sys.exit(1)
