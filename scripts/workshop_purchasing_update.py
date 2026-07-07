#!/usr/bin/env python3
"""Daily workshop purchasing tracker update from Odoo.
Runs at 2:00 PM daily via cron. Updates the tracker Excel with all SAMAYA WORKSHOP purchase orders.
"""
import xmlrpc.client
import ssl
from datetime import date, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
import os
import sys
import tempfile
import time

ODOO_URL = 'https://samayainv.odoo.com'
ODOO_DB = 'peerless-tech-samaya-18-0-18447146'
ODOO_USER = 'sultan@samayainvest.com'
OENV = os.path.expanduser('~/.config/samaya/odoo.env')
TARGET = "/Users/mohamedessa/Library/CloudStorage/OneDrive-SAMAYAINVESTMENT/Samaya/Orders/2026/0000 اداريات/شراء/ورشة المشتريات.xlsx"

def get_pw():
    with open(OENV) as f:
        for line in f:
            if 'ODOO_API_KEY' in line and '=' in line:
                return line.split('=', 1)[1].strip()
    return None

def connect():
    ctx = ssl._create_unverified_context()
    transport = xmlrpc.client.SafeTransport(context=ctx)
    common = xmlrpc.client.ServerProxy(ODOO_URL + '/xmlrpc/2/common', transport=transport)
    pw = get_pw()
    uid = common.authenticate(ODOO_DB, ODOO_USER, pw, {})
    models = xmlrpc.client.ServerProxy(ODOO_URL + '/xmlrpc/2/object', transport=transport)
    return uid, models

def fetch_pos(models, uid, pw):
    limit = 500
    pos = models.execute_kw(ODOO_DB, uid, pw, 'purchase.order', 'search_read',
        [[('user_id', '=', 155)]],
        {'fields': ['name', 'partner_id', 'partner_ref', 'amount_total', 'state',
                    'receipt_status', 'invoice_status', 'date_order'],
         'order': 'date_order desc',
         'limit': limit})
    return pos

def state_label(s):
    labels = {
        'draft': 'RFQ',
        'sent': 'RFQ Sent',
        'to approve': 'To Approve',
        'purchase': 'Purchase Order',
        'done': 'Locked',
        'cancel': 'Cancelled',
    }
    return labels.get(s, s)

def receipt_label(s):
    if not s:
        return ''
    labels = {
        'no': 'Not Received',
        'partial': 'Partially Received',
        'full': 'Received',
    }
    return labels.get(s, s)

def invoice_label(s):
    labels = {
        'no': 'Not Billed',
        'to invoice': 'Waiting Bills',
        'invoiced': 'Billed',
    }
    return labels.get(s, s)

def build_tracker(pos):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchasing Tracker"

    # Styles
    hdr_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    hdr_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )
    data_font = Font(name='Calibri', size=10)
    data_align = Alignment(vertical='center', wrap_text=True)
    date_fmt = 'DD-MM-YYYY'
    amt_fmt = '#,##0.00'

    # Title row
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value = f'ورشة المشتريات — Workshop Purchasing Tracker (Updated: {datetime.now().strftime("%d-%m-%Y %H:%M")})'
    title_cell.font = Font(name='Calibri', bold=True, size=14, color='1E293B')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Headers
    headers = [
        'PO Ref',
        'Vendor',
        'Vendor Reference',
        'Amount (SAR)',
        'Status',
        'Receipt Status',
        'Billing Status',
        'Date',
        'Notes/تحويل',
    ]
    col_widths = [12, 30, 40, 14, 14, 16, 14, 14, 30]

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(ci)].width = col_widths[ci-1]

    # Data rows
    row = 3
    amber_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
    green_fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
    red_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')

    for po in pos:
        name = po['name'] or ''
        partner = po['partner_id'][1] if po.get('partner_id') else ''
        partner_ref = po.get('partner_ref') or ''
        amount = po.get('amount_total') or 0.0
        state = state_label(po.get('state', ''))
        receipt = receipt_label(po.get('receipt_status'))
        invoice_st = invoice_label(po.get('invoice_status'))
        order_date = po.get('date_order', '')
        if order_date:
            try:
                dt = datetime.strptime(str(order_date)[:19], '%Y-%m-%d %H:%M:%S')
                order_date = dt.strftime('%d-%m-%Y')
            except:
                order_date = str(order_date)[:10]

        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=partner)
        ws.cell(row=row, column=3, value=partner_ref)
        amt_cell = ws.cell(row=row, column=4, value=float(amount))
        amt_cell.number_format = amt_fmt
        ws.cell(row=row, column=5, value=state)
        ws.cell(row=row, column=6, value=receipt)
        ws.cell(row=row, column=7, value=invoice_st)
        ws.cell(row=row, column=8, value=order_date)
        ws.cell(row=row, column=9, value='')

        # Apply row styling
        for ci in range(1, 10):
            cell = ws.cell(row=row, column=ci)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border
            if ci == 4:
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif ci in [5, 6, 7]:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Color coding
        if state in ('Purchase Order', 'Locked'):
            for ci in range(1, 10):
                ws.cell(row=row, column=ci).fill = green_fill
        elif state == 'RFQ':
            for ci in range(1, 10):
                ws.cell(row=row, column=ci).fill = amber_fill
        if receipt == 'Not Received' and state in ('Purchase Order', 'Locked'):
            pass  # keep green but could highlight

        row += 1

    # Freeze panes
    ws.freeze_panes = 'A3'

    # Auto-filter
    ws.auto_filter.ref = f'A2:I{row-1}'

    out = TARGET
    out_dir = os.path.dirname(out)
    os.makedirs(out_dir, exist_ok=True)

    # OneDrive can temporarily lock the target XLSX. Save to a local temp
    # file first, then atomically replace with retries to avoid cron failures.
    fd, tmp = tempfile.mkstemp(prefix='.workshop_purchasing_', suffix='.xlsx', dir=out_dir)
    os.close(fd)
    try:
        wb.save(tmp)
        last_error = None
        for attempt in range(5):
            try:
                os.replace(tmp, out)
                return out
            except OSError as exc:
                last_error = exc
                time.sleep(2 * (attempt + 1))
        # If OneDrive keeps the existing cloud placeholder inaccessible, keep
        # the generated tracker beside it instead of failing the cron job.
        fallback = os.path.join(out_dir, f"ورشة المشتريات_updated_{date.today().isoformat()}.xlsx")
        os.replace(tmp, fallback)
        print(f"⚠️  OneDrive target locked/unavailable, wrote fallback: {fallback}")
        return fallback
    finally:
        if os.path.exists(tmp):
            os.remove(tmp)

def main():
    uid, models = connect()
    pw = get_pw()
    pos = fetch_pos(models, uid, pw)
    path = build_tracker(pos)
    print(f"✅  Workshop Purchasing Tracker updated: {path}")
    print(f"   {len(pos)} purchase orders synced from Odoo")

if __name__ == '__main__':
    main()
