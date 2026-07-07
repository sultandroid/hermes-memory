#!/usr/bin/env python3
"""
Register Updater — ONLY appends rows to existing Excel registers.
NEVER creates new files. Scans project for existing xlsx files, finds the best
match for each register type, and appends data rows.

Usage:
  python3 register_generator.py <project_dir> [--dry-run]
  
  --dry-run: show what would be appended without modifying files
"""

import os, sys, json, shutil, tempfile
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

BIM_ROOT = os.path.expanduser("~/Library/CloudStorage/OneDrive-SAMAYAINVESTMENT/Samaya/Technical Office/Bim Unit")


def find_excel_files(project_dir):
    """Find all existing xlsx/xls files in a project."""
    project_path = os.path.join(BIM_ROOT, project_dir) if not os.path.isabs(project_dir) else project_dir
    files = []
    for dirpath, _, filenames in os.walk(project_path):
        for f in filenames:
            if f.endswith(('.xlsx', '.xls')):
                files.append(os.path.join(dirpath, f))
    return files


def read_headers(filepath):
    """Read headers from first data sheet."""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    sheet = wb.active
    headers = [str(cell.value or "").strip() for cell in sheet[1]] if sheet.max_row > 0 else []
    wb.close()
    return headers


def find_best_register(excel_files, register_type):
    """Find the best existing file for a register type by matching keywords."""
    keywords = {
        "submittal": ["submittal", "submission", "doc", "mar", "sdr"],
        "drawing": ["drawing", "dwg", "cad"],
        "rfi": ["rfi", "request for information", "query"],
        "transmittal": ["transmittal", "transmit"],
        "contract": ["contract", "agreement", "po", "purchase"],
        "invoice": ["invoice", "bill", "payment"],
        "material": ["material", "mar", "sample"],
        "meeting": ["meeting", "minutes", "mom"],
        "boq": ["boq", "bill of quantity", "pricing", "quantity"],
        "risk": ["risk", "hazard"],
        "change": ["change", "variation", "co"],
        "hse": ["hse", "safety", "health"],
        "ncr": ["ncr", "non-conformance", "nonconformance"],
        "subcontractor": ["subcontractor", "subcon", "vendor"],
        "si": ["si", "site instruction", "instruction"],
    }
    
    reg_keywords = keywords.get(register_type.lower(), [register_type.lower()])
    
    best_score = 0
    best_file = None
    
    for filepath in excel_files:
        fname = os.path.basename(filepath).lower()
        score = sum(1 for kw in reg_keywords if kw in fname)
        # Also check headers
        try:
            headers = read_headers(filepath)
            score += sum(1 for kw in reg_keywords for h in headers if kw in h.lower())
        except:
            pass
        if score > best_score:
            best_score = score
            best_file = filepath
    
    return best_file


def append_rows(filepath, new_rows, dry_run=False):
    """Append rows to existing Excel file, preserving formatting."""
    if dry_run:
        print(f"  [DRY RUN] Would append {len(new_rows)} rows to {os.path.basename(filepath)}")
        return True
    
    # Write to temp then atomic rename
    fd, temppath = tempfile.mkstemp(suffix=".xlsx", dir=os.path.dirname(filepath))
    os.close(fd)
    
    wb = load_workbook(filepath)
    sheet = wb.active
    
    # Find last row with data
    last_row = sheet.max_row
    
    for i, row_data in enumerate(new_rows):
        for j, value in enumerate(row_data, 1):
            cell = sheet.cell(row=last_row + 1 + i, column=j, value=value)
    
    wb.save(temppath)
    wb.close()
    
    shutil.move(temppath, filepath)
    print(f"  ✅ Appended {len(new_rows)} rows to {os.path.basename(filepath)}")
    return True


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(f"Usage: {sys.argv[0]} <project_dir> [--dry-run]")
        sys.exit(1)
    
    project_dir = sys.argv[1]
    dry_run = "--dry-run" in sys.argv
    
    existing = find_excel_files(project_dir)
    
    print(f"\n📋 Scanning {project_dir} for existing Excel files...")
    print(f"   Found {len(existing)} Excel files")
    
    for f in existing[:20]:
        rel = os.path.relpath(f, os.path.join(BIM_ROOT, project_dir))
        print(f"   📄 {rel}")
    if len(existing) > 20:
        print(f"   ... and {len(existing)-20} more")
    
    if not existing:
        print(f"\n⚠️  No existing Excel files found in {project_dir}")
        print(f"   Cannot append — no files to update.")
        print(f"   Ask the user if they want to create a new file.")
    
    print(f"\n{'='*60}")
    print(f"Mode: {'DRY RUN (no changes)' if dry_run else 'Ready to append'}")
    print(f"To append data: use Kimi or Claude Code to extract data and call append_rows()")
