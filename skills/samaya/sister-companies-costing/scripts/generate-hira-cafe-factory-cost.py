#!/usr/bin/env python3
"""
Generate Hira Cafe (Project 04) Factory Cost Details Excel files.

Template for per-project factory cost detail generators.
Copy this file and replace the data tuples + paths for each project.

Data structure:
  labour_data = [(trade, records, hours, cost), ...]  # 4-element tuples
  materials_data = [(po_ref, description, amount), ...]  # 3-element tuples
  other_data = [(description, amount), ...]  # 2-element tuples

Pitfall: tuple indices differ per data type!
  labour[3] = cost, materials[2] = amount, other[1] = amount
"""

import os
import shutil
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Styling constants ──────────────────────────────────────────────────
NAVY = "1F3864"
WHITE = "FFFFFF"
YELLOW = "FFD700"

header_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

total_font = Font(name="Calibri", bold=True, size=10, color="000000")
total_fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
total_align = Alignment(horizontal="center", vertical="center")

num_fmt = '#,##0.00'

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)

data_align = Alignment(horizontal="center", vertical="center")
data_align_left = Alignment(horizontal="left", vertical="center")


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def style_data_row(ws, row, cols, bold=False):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name="Calibri", bold=bold, size=10)
        cell.alignment = data_align if c != 2 else data_align_left
        cell.border = thin_border


def style_total_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = total_align
        cell.border = thin_border


def set_number_format(ws, row, col, fmt=num_fmt):
    ws.cell(row=row, column=col).number_format = fmt


# ═══════════════════════════════════════════════════════════════════════
# DATA — Replace these for each project
# ═══════════════════════════════════════════════════════════════════════

labour_data = [
    ("Welder", 68, 601.5, 9248),
    ("Painter", 72, 592, 8434),
    ("Carpenter", 31, 268.4, 2237),
    ("Labor", 27, 222.4, 1815),
    ("CNC Operator", 12, 106.5, 1041),
    ("Veneer Technician", 7, 66.5, 791),
    ("Supervisor", 3, 14.5, 194),
    ("3D Printers Operator", 2, 18.5, 137),
]

materials_data = [
    ("PO#416", "Paint", 25000),
    ("PO#609,623,891,894", "Upholstery", 17335),
    ("PO#418,465", "Electrical", 8900),
    ("PO#472", "Marble", 4700),
    ("PO#476", "Other", 3000),
    ("PO#454", "Metal/Steel", 930),
    ("Fleet/Transport", "Fleet/Transport", 1400),
]

other_data = [
    ("Reallocation (billboard moved to Ice Coffee 12)", -29549),
]

# ═══════════════════════════════════════════════════════════════════════
# BUILDERS
# ═══════════════════════════════════════════════════════════════════════


def build_labour(ws):
    ws.title = "Labour"
    headers = ["#", "Trade", "Records", "Hours", "Cost (SAR)"]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    for i, (trade, recs, hrs, cost) in enumerate(labour_data, 1):
        row = i + 1
        ws.append([i, trade, recs, hrs, cost])
        style_data_row(ws, row, len(headers))
        set_number_format(ws, row, 3, '#,##0')
        set_number_format(ws, row, 4, '#,##0.00')
        set_number_format(ws, row, 5, '#,##0.00')

    total_row = len(labour_data) + 2
    ws.append(["", "Total", sum(r[1] for r in labour_data),
               sum(r[2] for r in labour_data), sum(r[3] for r in labour_data)])
    style_total_row(ws, total_row, len(headers))
    set_number_format(ws, total_row, 3, '#,##0')
    set_number_format(ws, total_row, 4, '#,##0.00')
    set_number_format(ws, total_row, 5, '#,##0.00')

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 16


def build_materials(ws):
    ws.title = "Materials & POs"
    headers = ["PO#", "Description", "Amount (SAR)"]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    for po, desc, amt in materials_data:
        row = ws.max_row + 1
        ws.append([po, desc, amt])
        style_data_row(ws, row, len(headers))
        set_number_format(ws, row, 3, '#,##0.00')

    total_row = ws.max_row + 1
    ws.append(["", "Total", sum(r[2] for r in materials_data)])
    style_total_row(ws, total_row, len(headers))
    set_number_format(ws, total_row, 3, '#,##0.00')

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 16


def build_other(ws):
    ws.title = "Other Expenses"
    headers = ["#", "Description", "Amount (SAR)", "Notes"]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    for i, (desc, amt) in enumerate(other_data, 1):
        row = i + 1
        ws.append([i, desc, amt, ""])
        style_data_row(ws, row, len(headers))
        set_number_format(ws, row, 3, '#,##0.00')

    total_row = len(other_data) + 2
    ws.append(["", "Total", sum(r[1] for r in other_data), ""])
    style_total_row(ws, total_row, len(headers))
    set_number_format(ws, total_row, 3, '#,##0.00')

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 40


def build_summary(ws):
    ws.title = "Summary"
    headers = ["Category", "Amount (SAR)"]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    summary_data = [
        ("Labour", sum(r[3] for r in labour_data)),
        ("Materials", sum(r[2] for r in materials_data)),
        ("Other", sum(r[1] for r in other_data)),
    ]
    total_all = sum(r[1] for r in summary_data)

    for cat, amt in summary_data:
        row = ws.max_row + 1
        ws.append([cat, amt])
        style_data_row(ws, row, len(headers))
        set_number_format(ws, row, 2, '#,##0.00')

    total_row = ws.max_row + 1
    ws.append(["Total", total_all])
    style_total_row(ws, total_row, len(headers))
    set_number_format(ws, total_row, 2, '#,##0.00')

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 18


def build_gap_analysis(ws):
    ws.title = "Gap_Analysis"
    headers = ["Category", "Target (SAR)", "Actual (SAR)", "Gap (SAR)"]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    labour_total = sum(r[3] for r in labour_data)
    materials_total = sum(r[2] for r in materials_data)
    other_total = sum(r[1] for r in other_data)
    grand_total = labour_total + materials_total + other_total

    gap_data = [
        ("Labour", labour_total, labour_total, 0),
        ("Materials", materials_total, materials_total, 0),
        ("Other", other_total, other_total, 0),
        ("Total", grand_total, grand_total, 0),
    ]

    for cat, target, actual, gap in gap_data:
        row = ws.max_row + 1
        ws.append([cat, target, actual, gap])
        style_data_row(ws, row, len(headers))
        set_number_format(ws, row, 2, '#,##0.00')
        set_number_format(ws, row, 3, '#,##0.00')
        set_number_format(ws, row, 4, '#,##0.00')

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18


def create_full_workbook():
    wb = Workbook()
    build_labour(wb.active)
    build_materials(wb.create_sheet())
    build_other(wb.create_sheet())
    build_summary(wb.create_sheet())
    build_gap_analysis(wb.create_sheet())
    return wb


def create_clean_workbook():
    wb = Workbook()
    build_labour(wb.active)
    build_materials(wb.create_sheet())
    build_other(wb.create_sheet())
    return wb


# ═══════════════════════════════════════════════════════════════════════
# MAIN — Update paths for each project
# ═══════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    BASE = os.path.expanduser(
        "~/OneDrive - SAMAYA INVESTMENT/Reports/Sister_Companies"
    )

    # UPDATE THESE for each project:
    PROJECT_CODE = "04_Hira_Cafe"
    COMPANY_ORGANIZED = "Tiba_Gift_Company"  # from 00_Organized_13_Project_Factory_Reconciliation/
    CLIENT_FINAL = "Tiba Gift comp_"          # from _Final/

    organized_dir = os.path.join(
        BASE, "00_Organized_13_Project_Factory_Reconciliation",
        COMPANY_ORGANIZED, PROJECT_CODE
    )
    final_dir = os.path.join(
        BASE, "_Final", CLIENT_FINAL, PROJECT_CODE
    )

    os.makedirs(organized_dir, exist_ok=True)
    os.makedirs(final_dir, exist_ok=True)

    # Full version
    full_path = os.path.join(organized_dir, f"{PROJECT_CODE}_Factory_Cost_Details.xlsx")
    wb_full = create_full_workbook()
    wb_full.save(full_path)
    print(f"✅ Created: {full_path}")

    # Clean version
    clean_path = os.path.join(organized_dir, f"{PROJECT_CODE}_Factory_Cost_Details_Clean.xlsx")
    wb_clean = create_clean_workbook()
    wb_clean.save(clean_path)
    print(f"✅ Created: {clean_path}")

    # Copy to _Final
    for fname in [f"{PROJECT_CODE}_Factory_Cost_Details.xlsx",
                  f"{PROJECT_CODE}_Factory_Cost_Details_Clean.xlsx"]:
        src = os.path.join(organized_dir, fname)
        dst = os.path.join(final_dir, fname)
        shutil.copy2(src, dst)
        print(f"✅ Copied to _Final: {dst}")

    print(f"\n🎉 Done! Both files created and copied.")
