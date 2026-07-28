#!/usr/bin/env python3
"""Build Samaya Factory PO report — grouped by project (extracted from vendor reference).
Includes credit supplier rows per project group, PO Numbers column, total needed per project.
Uses fast bill check (first 3 invoices only) to avoid timeout."""
import os, sys, xmlrpc.client, ssl, re
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ENV_PATH = os.path.expanduser('~/.config/samaya/odoo.env')
env = {}
with open(ENV_PATH) as f:
    for line in f:
        line = line.strip()
        if not line or line.startswith('#') or '=' not in line: continue
        k, v = line.split('=', 1)
        env[k.strip()] = v.strip()

ctx = ssl._create_unverified_context()
common = xmlrpc.client.ServerProxy(f'{env["ODOO_URL"]}/xmlrpc/2/common', context=ctx)
uid = common.authenticate(env['ODOO_DB'], env['ODOO_USER'], env['ODOO_API_KEY'], {})
print(f'uid={uid}', file=sys.stderr)
models = xmlrpc.client.ServerProxy(f'{env["ODOO_URL"]}/xmlrpc/2/object', context=ctx)

# Fetch all Factory POs (purchase/done + draft)
d = [['state','in',['purchase','done','draft']], ['date_order','>=','2026-01-01']]
flds = ['name','partner_id','partner_ref','amount_total','date_order','state',
        'receipt_status','invoice_status','project_id','invoice_ids']
rows = models.execute_kw(env['ODOO_DB'], uid, env['ODOO_API_KEY'],
    'purchase.order', 'search_read', [d], {'fields': flds, 'limit': 2000})
factory_pos = [po for po in rows if po.get('project_id') and isinstance(po['project_id'], list)
               and len(po['project_id']) >= 1 and po['project_id'][0] == 244]
print(f'Factory POs: {len(factory_pos)}', file=sys.stderr)

cp_set = {'P01924','P01939','P01894','P01977'}
MADA_ID=2427; SABA_ID=5603

def bill_fast(po):
    inv_ids = po.get('invoice_ids') or []
    if not inv_ids: return 'no_bill', po.get('amount_total',0)
    rt = po.get('amount_total',0); ps = 'no_bill'
    for iid in inv_ids[:3]:
        try:
            b = models.execute_kw(env['ODOO_DB'], uid, env['ODOO_API_KEY'],
                'account.move', 'read', [iid],
                {'fields': ['amount_total','amount_residual','payment_state','state']})
            if b and len(b)>0:
                b=b[0]; r=b.get('amount_residual',b.get('amount_total',0))
                if b.get('state')=='posted' and abs(r)<0.01: return 'paid',0
                elif b.get('state')=='posted' and r<b.get('amount_total',0): rt=r; ps='partial'
                elif b.get('state')=='draft': rt=r; ps='draft_bill'
        except: pass
    return ps, rt

def classify(ref):
    ref=(ref or '').strip()
    if not ref: return 'غير محدد'
    pats = [
        (r'Jalal.*(?:Jabal Omer|جبل عمر)', 'Jalal & Jamal'),
        (r'Maalim.*(?:Jabal Omer|جبل عمر)', 'Maalim Al-Haramein'),
        (r'متاجر الغمامة', 'متاجر الغمامة'),
        (r'متجر الهدايا.*(?:معالم الحرمين|جبل عمر)', 'متجر الهدايا - معالم الحرمين'),
        (r'متحف عسير', 'متحف عسير الإقليمي'),
        (r'متحف القرآن|القران الكريم', 'متحف القرآن الكريم'),
        (r'متحف خير الخلق', 'متحف خير الخلق'),
        (r'متحف الغمامة', 'متحف الغمامة'),
        (r'متحف معالم المسجد|معالم المسجد', 'متحف معالم المسجد الحرام'),
        (r'زمزم|Zamzam', 'Zamzam'),
        (r'غار حراء|المركز الإعلامي|حراء', 'المركز الإعلامي - حراء'),
        (r'جبل عمر', 'جبل عمر (عام)'),
        (r'المصنع\s*-?\s*(?:كشف|مستلزمات|مواد|طلب|جلفزات|مصاريف)?', 'المصنع (تشغيلي)'),
        (r'بدل اعاشة|مصاريف تشغيلية|مصروفات الإعاشة', 'مصاريف تشغيلية'),
        (r'مدفوع من العهده|مدفوع م العهده|مدفوعه', 'عهدة إبراهيم'),
        (r'Expenses Statement', 'مصاريف تشغيلية'),
        (r'Outsorce|عمالة', 'عمالة خارجية'),
    ]
    for p,l in pats:
        if re.search(p,ref,re.UNICODE): return l
    return 'أخرى'

results = []
for po in factory_pos:
    pn = po.get('partner_id',['',''])[1] if po.get('partner_id') else ''
    pv = po.get('partner_id',[0])[0] if po.get('partner_id') else 0
    nm=po['name']; rf=po.get('partner_ref','') or ''
    if pv in {MADA_ID,SABA_ID}: continue
    ps,rt = bill_fast(po)
    pd=False; src='unpaid'
    if ps=='paid': pd=True; src='bill_paid'
    elif nm in cp_set: pd=True; src='chatter_paid'
    results.append({'po':nm,'vendor':pn,'amount':po.get('amount_total',0),
        'date':str(po.get('date_order',''))[:10],'state':po.get('state',''),
        'paid':pd,'pay_source':src,'project_group':classify(rf),'ref':rf[:80]})

groups={}
for r in results:
    g=r['project_group']
    if g not in groups: groups[g]={'unpaid':[],'bill_paid':[],'chatter_paid':[]}
    if r['pay_source']=='bill_paid': groups[g]['bill_paid'].append(r)
    elif r['pay_source']=='chatter_paid': groups[g]['chatter_paid'].append(r)
    else: groups[g]['unpaid'].append(r)

cr = [po for po in factory_pos if (po.get('partner_id',[0])[0] if po.get('partner_id') else 0) in {MADA_ID,SABA_ID}]
cby={}
for po in cr:
    pv=po.get('partner_id',[0])[0]; g=classify(po.get('partner_ref','') or '')
    if g not in cby: cby[g]={'mada_t':0,'saba_t':0,'mada_c':0,'saba_c':0}
    a=po.get('amount_total',0)
    if pv==MADA_ID: cby[g]['mada_t']+=a; cby[g]['mada_c']+=1
    elif pv==SABA_ID: cby[g]['saba_t']+=a; cby[g]['saba_c']+=1

sg = sorted(groups.items(), key=lambda x: sum(r['amount'] for r in x[1]['unpaid']), reverse=True)

wb=Workbook()
nvy=PatternFill('solid',fgColor='1F3864'); gld=PatternFill('solid',fgColor='C9A84C')
lg=PatternFill('solid',fgColor='F2F2F2'); wf=PatternFill('solid',fgColor='FFFFFF')
gf=PatternFill('solid',fgColor='C6EFCE'); rf=PatternFill('solid',fgColor='FFC7CE')
yf=PatternFill('solid',fgColor='FFEB9C'); bf=PatternFill('solid',fgColor='D6E4F0')
pf=PatternFill('solid',fgColor='E8D4F0')
hf=Font(name='Calibri',bold=True,color='FFFFFF',size=11)
tf=Font(name='Calibri',bold=True,color='1F3864',size=14)
bfont=Font(name='Calibri',size=10); boldf=Font(name='Calibri',bold=True,size=10)
totf=Font(name='Calibri',bold=True,color='C9A84C',size=11)
tb=Border(left=Side(style='thin',color='D0D0D0'),right=Side(style='thin',color='D0D0D0'),
          top=Side(style='thin',color='D0D0D0'),bottom=Side(style='thin',color='D0D0D0'))
def hdr(ws,r,n):
    for c in range(1,n+1):
        cl=ws.cell(row=r,column=c); cl.fill=nvy; cl.font=hf
        cl.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); cl.border=tb
def sc(ws,r,c,v,font=None,fill=None,align=None):
    cl=ws.cell(row=r,column=c,value=v); cl.font=font or bfont; cl.fill=fill or wf
    cl.alignment=align or Alignment(vertical='center',wrap_text=True); cl.border=tb; return cl

# Sheet 1: By Project
ws1=wb.active; ws1.title='By Project'
ws1.merge_cells('A1:I1')
ws1['A1'].value='SAMAYA FACTORY — CASHOUT BY PROJECT (from Vendor Reference)'
ws1['A1'].font=tf; ws1['A1'].alignment=Alignment(horizontal='center')
ws1.merge_cells('A2:I2')
ws1['A2'].value=f'Prepared {date.today().strftime("%d %b %Y")} | All amounts in SAR'
ws1['A2'].font=Font(name='Calibri',color='64748B',size=10); ws1['A2'].alignment=Alignment(horizontal='center')
hds=['Project (from Vendor Ref)','Unpaid POs','PO Numbers','Unpaid (SAR)','Bill-Paid (SAR)',
     'Chatter-Paid (SAR)','Credit Supp. (SAR)','Total Needed (SAR)']
row=4
for i,h in enumerate(hds,1): ws1.cell(row=row,column=i,value=h)
hdr(ws1,row,8)
row=5; gu=0; gb=0; gc=0; gcr=0
for gn,gd in sg:
    ut=round(sum(r['amount'] for r in gd['unpaid']),2)
    bt=round(sum(r['amount'] for r in gd['bill_paid']),2)
    ct=round(sum(r['amount'] for r in gd['chatter_paid']),2)
    cp=cby.get(gn,{'mada_t':0,'saba_t':0})
    crt=round(cp['mada_t']+cp['saba_t'],2)
    tn=round(ut+crt,2)
    if ut==0 and bt==0 and ct==0 and crt==0: continue
    al=lg if row%2==0 else wf
    pnums=', '.join(r['po'] for r in gd['unpaid'])
    sc(ws1,row,1,gn,boldf if ut>0 else bfont,al)
    sc(ws1,row,2,len(gd['unpaid']),bfont,al,Alignment(horizontal='center'))
    sc(ws1,row,3,pnums,bfont,al)
    sc(ws1,row,4,ut,boldf if ut>0 else bfont,rf if ut>0 else al,Alignment(horizontal='right'))
    sc(ws1,row,5,bt,bfont,gf if bt>0 else al,Alignment(horizontal='right'))
    sc(ws1,row,6,ct,bfont,bf if ct>0 else al,Alignment(horizontal='right'))
    sc(ws1,row,7,crt,boldf if crt>0 else bfont,pf if crt>0 else al,Alignment(horizontal='right'))
    sc(ws1,row,8,tn,boldf,gld if tn>0 else al,Alignment(horizontal='right'))
    gu+=ut; gb+=bt; gc+=ct; gcr+=crt; row+=1

tc=sum(len(g['unpaid']) for g in groups.values())
sc(ws1,row,1,'GRAND TOTAL',totf,gld); sc(ws1,row,2,tc,totf,gld,Alignment(horizontal='center'))
for c in range(3,9): sc(ws1,row,c,'',totf,gld)
cws=[42,12,55,16,16,16,20,18]
for i,w in enumerate(cws,1): ws1.column_dimensions[get_column_letter(i)].width=w

# Sheet 2: Detail
ws2=wb.create_sheet('Detail')
ws2.merge_cells('A1:I1')
ws2['A1'].value='SAMAYA FACTORY — PO Detail by Project Group'
ws2['A1'].font=tf; ws2['A1'].alignment=Alignment(horizontal='center')
row=3
dh=['Project','PO #','Vendor','Amount (SAR)','Date','State','Payment','Credit Supp.','Vendor Reference']
for i,h in enumerate(dh,1): ws2.cell(row=row,column=i,value=h)
hdr(ws2,row,9); row=4
for gn,gd in sg:
    ai=sorted(gd['unpaid']+gd['bill_paid']+gd['chatter_paid'],key=lambda r: (r['pay_source']!='unpaid',r['date']))
    if not ai: continue
    ut=sum(r['amount'] for r in gd['unpaid'])
    ws2.merge_cells(start_row=row,start_column=1,end_row=row,end_column=9)
    sc(ws2,row,1,f"{gn} — {len(gd['unpaid'])} unpaid = {ut:,.2f} SAR",boldf,nvy,Alignment(horizontal='left'))
    for c in range(2,10): ws2.cell(row=row,column=c).fill=nvy
    row+=1
    for i,r in enumerate(ai):
        al=lg if i%2==0 else wf
        pf2=gf if r['pay_source']=='bill_paid' else (bf if r['pay_source']=='chatter_paid' else (yf if r['state']=='draft' else al))
        sc(ws2,row,1,'',bfont,al); sc(ws2,row,2,r['po'],bfont,pf2)
        sc(ws2,row,3,r['vendor'],bfont,pf2); sc(ws2,row,4,round(r['amount'],2),bfont,pf2,Alignment(horizontal='right'))
        sc(ws2,row,5,r['date'],bfont,pf2,Alignment(horizontal='center'))
        sc(ws2,row,6,r['state'],bfont,pf2,Alignment(horizontal='center'))
        sc(ws2,row,7,r['pay_source'],bfont,pf2,Alignment(horizontal='center'))
        sc(ws2,row,8,'',bfont,pf2); sc(ws2,row,9,r['ref'],bfont,pf2)
        row+=1
    cp=cby.get(gn,{'mada_t':0,'saba_t':0,'mada_c':0,'saba_c':0})
    if cp.get('mada_c',0)>0:
        sc(ws2,row,3,'Credit: مؤسسة مدى الجزيرة',boldf,pf)
        sc(ws2,row,4,round(cp['mada_t'],2),boldf,pf,Alignment(horizontal='right'))
        sc(ws2,row,7,f'{cp["mada_c"]} POs',bfont,pf,Alignment(horizontal='center'))
        row+=1
    if cp.get('saba_c',0)>0:
        sc(ws2,row,3,'Credit: صبا نجد',boldf,pf)
        sc(ws2,row,4,round(cp['saba_t'],2),boldf,pf,Alignment(horizontal='right'))
        sc(ws2,row,7,f'{cp["saba_c"]} POs',bfont,pf,Alignment(horizontal='center'))
        row+=1
    row+=1

cw2=[42,14,35,16,12,10,14,14,60]
for i,w in enumerate(cw2,1): ws2.column_dimensions[get_column_letter(i)].width=w

out='/Users/mohamedessa/Library/CloudStorage/OneDrive-SAMAYAINVESTMENT/Samaya/Orders/2026/0000 اداريات/00 تقارير الاعمال/Samaya_Factory_PO_By_Project.xlsx'
wb.save(out)
print(f'Saved: {out}', file=sys.stderr)
print(f'Grand total needed: {gu+gcr:,.2f} SAR (unpaid: {gu:,.2f} + credit: {gcr:,.2f})')
