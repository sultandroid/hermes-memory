#!/usr/bin/env python3
"""Update workshop purchasing tracker with latest Odoo payment data."""
import os, sys, xmlrpc.client, ssl, re
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

ENV_PATH = os.path.expanduser('~/.config/samaya/odoo.env')
env = {}
with open(ENV_PATH) as f:
    for line in f:
        line = line.strip()
        if not line or line.startswith('#') or '=' not in line: continue
        k, v = line.split('=', 1)
        env[k.strip()] = v.strip()

ctx = ssl._create_unverified_context()
common = xmlrpc.client.ServerProxy(f'{env["ODOO_URL"]}/xmlrpc/2/common', context=ctx)
uid = common.authenticate(env['ODOO_DB'], env['ODOO_USER'], env['ODOO_API_KEY'], {})
print(f'uid={uid}', file=sys.stderr)
models = xmlrpc.client.ServerProxy(f'{env["ODOO_URL"]}/xmlrpc/2/object', context=ctx)

domain = [['state','in',['purchase','done']], ['date_order','>=','2026-01-01']]
fields = ['name','partner_id','amount_total','date_order','state','receipt_status','invoice_status','project_id','invoice_ids']
rows = models.execute_kw(env['ODOO_DB'], uid, env['ODOO_API_KEY'],
    'purchase.order', 'search_read', [domain], {'fields': fields, 'limit': 200})

po_lookup = {}
for po in rows:
    inv_ids = po.get('invoice_ids') or []
    pname = po.get('partner_id', ['',''])[1] if po.get('partner_id') else ''
    pay_state = 'no_bill'; paid = False; bill_info = []
    if inv_ids:
        for inv_id in inv_ids:
            try:
                b = models.execute_kw(env['ODOO_DB'], uid, env['ODOO_API_KEY'],
                    'account.move', 'read', [inv_id], {'fields': ['name','amount_total','amount_residual','payment_state','state']})
                if b and len(b) > 0:
                    b = b[0]; res = b.get('amount_residual', b.get('amount_total', 0))
                    bill_info.append(f"{b.get('name','')}={res:.0f}/{b.get('amount_total',0):.0f}")
                    if b.get('state') == 'posted' and abs(res) < 0.01: paid = True; pay_state = 'paid'
                    elif b.get('state') == 'posted' and res < b.get('amount_total',0): pay_state = 'partial'
                    elif b.get('state') == 'draft': pay_state = 'draft_bill'
            except: pass
    po_lookup[po['name']] = {'vendor': pname, 'amount': po.get('amount_total',0), 'state': po.get('state',''), 'receipt': po.get('receipt_status',''), 'inv_status': po.get('invoice_status',''), 'paid': paid, 'pay_state': pay_state, 'bills': '; '.join(bill_info)}

tracker_path = '/Users/mohamedessa/Library/CloudStorage/OneDrive-SAMAYAINVESTMENT/Samaya/Orders/2026/0000 اداريات/شراء/ورشة المشتريات.xlsx'
wb = load_workbook(tracker_path)
ws = wb.active

hdr_col = 10; hdr_col2 = 11
ws.cell(row=3, column=hdr_col, value='Odoo Pay State').font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
ws.cell(row=3, column=hdr_col).fill = PatternFill('solid', fgColor='1F3864')
ws.cell(row=3, column=hdr_col).alignment = Alignment(horizontal='center', wrap_text=True)
ws.cell(row=3, column=hdr_col2, value='Odoo Bills').font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
ws.cell(row=3, column=hdr_col2).fill = PatternFill('solid', fgColor='1F3864')
ws.cell(row=3, column=hdr_col2).alignment = Alignment(horizontal='center', wrap_text=True)

gf = PatternFill('solid', fgColor='C6EFCE'); yf = PatternFill('solid', fgColor='FFEB9C'); rf = PatternFill('solid', fgColor='FFC7CE')
updated = 0
for row in range(4, ws.max_row + 1):
    po_ref = ws.cell(row=row, column=1).value
    if not po_ref or po_ref == '': continue
    po_ref = str(po_ref).strip()
    if po_ref in po_lookup:
        info = po_lookup[po_ref]
        ws.cell(row=row, column=hdr_col, value=info['pay_state'])
        ws.cell(row=row, column=hdr_col2, value=info['bills'])
        if info['paid']: ws.cell(row=row, column=hdr_col).fill = gf
        elif info['pay_state'] == 'draft_bill': ws.cell(row=row, column=hdr_col).fill = yf
        elif info['pay_state'] == 'no_bill': ws.cell(row=row, column=hdr_col).fill = rf
        updated += 1
    else:
        ws.cell(row=row, column=hdr_col, value='not_in_odoo')
        ws.cell(row=row, column=hdr_col).fill = PatternFill('solid', fgColor='D9D9D9')

ws.cell(row=1, column=1).value = f'ورشة المشتريات — Workshop Purchasing Tracker (Updated: {date.today().strftime("%d-%m-%Y")})'
wb.save(tracker_path)
print(f'Updated {updated} POs in tracker', file=sys.stderr)
print(f'Done. {updated} POs updated with Odoo payment data.')
