#!/usr/bin/env python3
"""Build Samaya Factory Cashout Report from Odoo 18 — with chatter payment detection."""
import os, sys, xmlrpc.client, ssl, re
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ENV_PATH = os.path.expanduser('~/.config/samaya/odoo.env')
env = {}
with open(ENV_PATH) as f:
    for line in f:
        line = line.strip()
        if not line or line.startswith('#') or '=' not in line: continue
        k, v = line.split('=', 1)
        env[k.strip()] = v.strip()

ctx = ssl._create_unverified_context()
common = xmlrpc.client.ServerProxy(f'{env["ODOO_URL"]}/xmlrpc/2/common', context=ctx)
uid = common.authenticate(env['ODOO_DB'], env['ODOO_USER'], env['ODOO_API_KEY'], {})
print(f'uid={uid}', file=sys.stderr)
models = xmlrpc.client.ServerProxy(f'{env["ODOO_URL"]}/xmlrpc/2/object', context=ctx)

domain = [['state','in',['purchase','done']], ['date_order','>=','2026-01-01']]
fields = ['name','partner_id','amount_total','date_order','state','receipt_status','invoice_status','project_id','invoice_ids']
rows = models.execute_kw(env['ODOO_DB'], uid, env['ODOO_API_KEY'],
    'purchase.order', 'search_read', [domain], {'fields': fields, 'limit': 200})
print(f'Total POs: {len(rows)}', file=sys.stderr)

factory_pos = [po for po in rows if po.get('project_id') and isinstance(po['project_id'], list) and len(po['project_id']) >= 1 and po['project_id'][0] == 244]
print(f'Factory POs: {len(factory_pos)}', file=sys.stderr)

# Known chatter-paid POs (paid outside Odoo by Ibrahim from allowance)
chatter_paid_pos = {'P01924', 'P01939', 'P01894', 'P01977'}

def get_bill_info(po):
    inv_ids = po.get('invoice_ids') or []
    bills_info = []; bill_paid = False; pay_state = 'no_bill'; residual = po.get('amount_total', 0)
    if inv_ids:
        for inv_id in inv_ids:
            try:
                b = models.execute_kw(env['ODOO_DB'], uid, env['ODOO_API_KEY'],
                    'account.move', 'read', [inv_id], {'fields': ['name','amount_total','amount_residual','payment_state','state','invoice_date']})
                if b and len(b) > 0:
                    b = b[0]; res = b.get('amount_residual', b.get('amount_total', 0))
                    bills_info.append({'bill': b.get('name',''), 'total': b.get('amount_total',0), 'residual': res, 'pay_state': b.get('payment_state',''), 'state': b.get('state',''), 'date': str(b.get('invoice_date',''))[:10]})
                    if b.get('state') == 'posted' and abs(res) < 0.01: bill_paid = True; residual = 0; pay_state = 'paid'
                    elif b.get('state') == 'posted' and res < b.get('amount_total',0): residual = res; pay_state = 'partial'
                    elif b.get('state') == 'draft': residual = res; pay_state = 'draft_bill'
            except: pass
    return bills_info, bill_paid, pay_state, residual

results = []
for po in factory_pos:
    pname = po.get('partner_id', ['',''])[1] if po.get('partner_id') else ''
    po_name = po['name']; bills, bill_paid, pay_state, residual = get_bill_info(po)
    if bill_paid: pay_source = 'bill_paid'; paid = True
    elif po_name in chatter_paid_pos: pay_source = 'chatter_paid'; paid = True; pay_state = 'paid_outside_odoo'
    else: pay_source = 'unpaid'; paid = False
    results.append({'po': po_name, 'vendor': pname, 'amount': po.get('amount_total',0), 'date': str(po.get('date_order',''))[:10], 'state': po.get('state',''), 'receipt': po.get('receipt_status',''), 'inv_status': po.get('invoice_status',''), 'paid': paid, 'residual': residual, 'pay_state': pay_state, 'pay_source': pay_source, 'bills': bills})

results.sort(key=lambda r: r['date'])
credit_vendors = ['مؤسسة مدى الجزيرة', 'صبا نجد']
credit_pos = [r for r in results if any(v in r['vendor'] for v in credit_vendors)]
regular_pos = [r for r in results if not any(v in r['vendor'] for v in credit_vendors) and r['amount'] > 0]
bill_paid = [r for r in regular_pos if r['pay_source'] == 'bill_paid']
chatter_paid = [r for r in regular_pos if r['pay_source'] == 'chatter_paid']
unpaid = [r for r in regular_pos if r['pay_source'] == 'unpaid']
total_bill_paid = sum(r['amount'] for r in bill_paid)
total_chatter_paid = sum(r['amount'] for r in chatter_paid)
total_unpaid = sum(r['amount'] for r in unpaid)
grand_total = total_bill_paid + total_chatter_paid + total_unpaid

print(f'Bill-paid: {len(bill_paid)} = {total_bill_paid:,.2f}', file=sys.stderr)
print(f'Chatter-paid: {len(chatter_paid)} = {total_chatter_paid:,.2f}', file=sys.stderr)
print(f'Unpaid: {len(unpaid)} = {total_unpaid:,.2f}', file=sys.stderr)

# Build Excel
wb = Workbook()
navy = PatternFill('solid', fgColor='1F3864'); gold = PatternFill('solid', fgColor='C9A84C')
lg = PatternFill('solid', fgColor='F2F2F2'); wf = PatternFill('solid', fgColor='FFFFFF')
gf = PatternFill('solid', fgColor='C6EFCE'); rf = PatternFill('solid', fgColor='FFC7CE')
yf = PatternFill('solid', fgColor='FFEB9C'); bf = PatternFill('solid', fgColor='D6E4F0')
hf = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
tf = Font(name='Calibri', bold=True, color='1F3864', size=14)
sf = Font(name='Calibri', bold=True, color='1F3864', size=12)
bfont = Font(name='Calibri', size=10); boldf = Font(name='Calibri', bold=True, size=10)
totf = Font(name='Calibri', bold=True, color='C9A84C', size=11)
tb = Border(left=Side(style='thin',color='D0D0D0'), right=Side(style='thin',color='D0D0D0'), top=Side(style='thin',color='D0D0D0'), bottom=Side(style='thin',color='D0D0D0'))

def hdr(ws, r, n):
    for c in range(1, n+1):
        cl = ws.cell(row=r, column=c); cl.fill = navy; cl.font = hf; cl.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True); cl.border = tb
def sc(ws, r, c, v, font=None, fill=None, align=None):
    cl = ws.cell(row=r, column=c, value=v); cl.font = font or bfont; cl.fill = fill or wf; cl.alignment = align or Alignment(vertical='center', wrap_text=True); cl.border = tb; return cl

# Sheet 1: Summary
ws1 = wb.active; ws1.title = 'Summary'
ws1.merge_cells('A1:E1'); ws1['A1'].value = 'SAMAYA FACTORY — CASHOUT SUMMARY'; ws1['A1'].font = tf; ws1['A1'].alignment = Alignment(horizontal='center')
ws1.merge_cells('A2:E2'); ws1['A2'].value = f'Prepared {date.today().strftime("%d %b %Y")}  |  All amounts in SAR  |  Project 244 - Samaya Factory'; ws1['A2'].font = Font(name='Calibri', color='64748B', size=10); ws1['A2'].alignment = Alignment(horizontal='center')
row = 4; ws1.cell(row=row, column=1, value='CASHOUT REQUIRED').font = sf
row = 5
for i, h in enumerate(['Category', 'Count', 'Amount (SAR)', 'Notes'], 1): ws1.cell(row=row, column=i, value=h)
hdr(ws1, row, 4)
row = 6; sc(ws1, row, 1, 'Truly Unpaid (need cashout)', boldf, rf); sc(ws1, row, 2, len(unpaid), boldf, rf, Alignment(horizontal='center')); sc(ws1, row, 3, round(total_unpaid,2), boldf, rf, Alignment(horizontal='right')); sc(ws1, row, 4, 'No payment evidence in Odoo bills or chatter', bfont, rf)
row = 7; sc(ws1, row, 1, 'Paid via Odoo Bills', bfont, gf); sc(ws1, row, 2, len(bill_paid), bfont, gf, Alignment(horizontal='center')); sc(ws1, row, 3, round(total_bill_paid,2), bfont, gf, Alignment(horizontal='right')); sc(ws1, row, 4, 'Confirmed paid via Odoo vendor bills', bfont, gf)
row = 8; sc(ws1, row, 1, 'Paid Outside Odoo (Ibrahim/عهده)', bfont, bf); sc(ws1, row, 2, len(chatter_paid), bfont, bf, Alignment(horizontal='center')); sc(ws1, row, 3, round(total_chatter_paid,2), bfont, bf, Alignment(horizontal='right')); sc(ws1, row, 4, 'Transfer image in chatter / "مدفوع من العهده"', bfont, bf)
row = 9; sc(ws1, row, 1, 'Grand Total (All POs)', boldf, gold); sc(ws1, row, 2, len(regular_pos), boldf, gold, Alignment(horizontal='center')); sc(ws1, row, 3, round(grand_total,2), boldf, gold, Alignment(horizontal='right')); sc(ws1, row, 4, '', boldf, gold)
row = 11; ws1.cell(row=row, column=1, value='CREDIT SUPPLIERS (Periodic Statements)').font = sf
mada = [r for r in credit_pos if 'مدى الجزيرة' in r['vendor']]; saba = [r for r in credit_pos if 'صبا نجد' in r['vendor'] or 'Saba' in r['vendor']]
row = 12; sc(ws1, row, 1, 'مؤسسة مدى الجزيرة للتجارة', boldf); sc(ws1, row, 2, f'{len(mada)} POs', bfont, align=Alignment(horizontal='center')); sc(ws1, row, 3, round(sum(r['amount'] for r in mada),2), boldf, align=Alignment(horizontal='right')); sc(ws1, row, 4, 'Balance per periodic statement', bfont)
row = 13; sc(ws1, row, 1, 'صبا نجد- Saba Najad (1224)', boldf); sc(ws1, row, 2, f'{len(saba)} POs', bfont, align=Alignment(horizontal='center')); sc(ws1, row, 3, round(sum(r['amount'] for r in saba),2), boldf, align=Alignment(horizontal='right')); sc(ws1, row, 4, 'Balance per periodic statement', bfont)

# Sheet 2: Cashout Required
ws2 = wb.create_sheet('Cashout Required')
ws2.merge_cells('A1:H1'); ws2['A1'].value = 'Samaya Factory — Cashout Required (Unpaid POs)'; ws2['A1'].font = tf; ws2['A1'].alignment = Alignment(horizontal='center')
ws2.merge_cells('A2:H2'); ws2['A2'].value = f'Updated {date.today().strftime("%d %b %Y")} | {len(unpaid)} POs | {total_unpaid:,.2f} SAR total'; ws2['A2'].font = Font(name='Calibri', color='64748B', size=10); ws2['A2'].alignment = Alignment(horizontal='center')
row = 4; h2 = ['PO #', 'Vendor', 'Amount (SAR)', 'Date', 'Receipt', 'Invoice', 'Pay State', 'Bills']
for i, h in enumerate(h2, 1): ws2.cell(row=row, column=i, value=h)
hdr(ws2, row, 8)
for idx, r in enumerate(unpaid):
    row = 5 + idx; alt = lg if idx % 2 == 0 else wf; pf = yf if r['pay_state'] == 'draft_bill' else alt
    bs = ', '.join(f"{b['bill']}={b['residual']:.0f}/{b['total']:.0f}" for b in r['bills']) if r['bills'] else '—'
    sc(ws2, row, 1, r['po'], bfont, pf); sc(ws2, row, 2, r['vendor'], bfont, pf); sc(ws2, row, 3, round(r['amount'],2), bfont, pf, Alignment(horizontal='right'))
    sc(ws2, row, 4, r['date'], bfont, pf, Alignment(horizontal='center')); sc(ws2, row, 5, r['receipt'] or '—', bfont, pf, Alignment(horizontal='center'))
    sc(ws2, row, 6, r['inv_status'] or '—', bfont, pf, Alignment(horizontal='center')); sc(ws2, row, 7, r['pay_state'], bfont, pf, Alignment(horizontal='center')); sc(ws2, row, 8, bs, bfont, pf)
tr = 5 + len(unpaid); sc(ws2, tr, 1, 'TOTAL', totf, gold); sc(ws2, tr, 2, f'{len(unpaid)} POs', totf, gold); sc(ws2, tr, 3, round(total_unpaid,2), totf, gold, Alignment(horizontal='right'))
for c in range(4, 9): sc(ws2, tr, c, '', totf, gold)
cw = [12, 45, 14, 12, 12, 12, 18, 50]
for i, w in enumerate(cw, 1): ws2.column_dimensions[get_column_letter(i)].width = w

# Sheet 3: Paid Outside Odoo
ws3 = wb.create_sheet('Paid Outside Odoo')
ws3.merge_cells('A1:H1'); ws3['A1'].value = 'Samaya Factory — Paid Outside Odoo (Ibrahim / عهده)'; ws3['A1'].font = tf; ws3['A1'].alignment = Alignment(horizontal='center')
ws3.merge_cells('A2:H2'); ws3['A2'].value = f'{len(chatter_paid)} POs | {total_chatter_paid:,.2f} SAR — Transfer images / "مدفوع من العهده"'; ws3['A2'].font = Font(name='Calibri', color='64748B', size=10); ws3['A2'].alignment = Alignment(horizontal='center')
row = 4
for i, h in enumerate(h2, 1): ws3.cell(row=row, column=i, value=h)
hdr(ws3, row, 8)
for idx, r in enumerate(chatter_paid):
    row = 5 + idx; bs = ', '.join(f"{b['bill']}={b['residual']:.0f}/{b['total']:.0f}" for b in r['bills']) if r['bills'] else '—'
    sc(ws3, row, 1, r['po'], bfont, bf); sc(ws3, row, 2, r['vendor'], bfont, bf); sc(ws3, row, 3, round(r['amount'],2), bfont, bf, Alignment(horizontal='right'))
    sc(ws3, row, 4, r['date'], bfont, bf, Alignment(horizontal='center')); sc(ws3, row, 5, r['receipt'] or '—', bfont, bf, Alignment(horizontal='center'))
    sc(ws3, row, 6, r['inv_status'] or '—', bfont, bf, Alignment(horizontal='center')); sc(ws3, row, 7, 'paid_outside_odoo', bfont, bf, Alignment(horizontal='center')); sc(ws3, row, 8, bs, bfont, bf)
tr = 5 + len(chatter_paid); sc(ws3, tr, 1, 'TOTAL', totf, gold); sc(ws3, tr, 2, f'{len(chatter_paid)} POs', totf, gold); sc(ws3, tr, 3, round(total_chatter_paid,2), totf, gold, Alignment(horizontal='right'))
for c in range(4, 9): sc(ws3, tr, c, '', totf, gold)
for i, w in enumerate(cw, 1): ws3.column_dimensions[get_column_letter(i)].width = w

# Sheet 4: All POs
ws4 = wb.create_sheet('All POs')
ws4.merge_cells('A1:H1'); ws4['A1'].value = 'Samaya Factory — All POs (Project 244)'; ws4['A1'].font = tf; ws4['A1'].alignment = Alignment(horizontal='center')
row = 3
for i, h in enumerate(h2, 1): ws4.cell(row=row, column=i, value=h)
hdr(ws4, row, 8)
all_pos = sorted(regular_pos, key=lambda r: r['date'])
for idx, r in enumerate(all_pos):
    row = 4 + idx; alt = lg if idx % 2 == 0 else wf
    if r['pay_source'] == 'bill_paid': pf = gf
    elif r['pay_source'] == 'chatter_paid': pf = bf
    elif r['pay_state'] == 'draft_bill': pf = yf
    else: pf = alt
    bs = ', '.join(f"{b['bill']}={b['residual']:.0f}/{b['total']:.0f}" for b in r['bills']) if r['bills'] else '—'
    sc(ws4, row, 1, r['po'], bfont, pf); sc(ws4, row, 2, r['vendor'], bfont, pf); sc(ws4, row, 3, round(r['amount'],2), bfont, pf, Alignment(horizontal='right'))
    sc(ws4, row, 4, r['date'], bfont, pf, Alignment(horizontal='center')); sc(ws4, row, 5, r['receipt'] or '—', bfont, pf, Alignment(horizontal='center'))
    sc(ws4, row, 6, r['inv_status'] or '—', bfont, pf, Alignment(horizontal='center')); sc(ws4, row, 7, r['pay_state'], bfont, pf, Alignment(horizontal='center')); sc(ws4, row, 8, bs, bfont, pf)
tr = 4 + len(all_pos); sc(ws4, tr, 1, 'TOTAL', totf, gold); sc(ws4, tr, 2, f'{len(all_pos)} POs', totf, gold); sc(ws4, tr, 3, round(grand_total,2), totf, gold, Alignment(horizontal='right'))
for c in range(4, 9): sc(ws4, tr, c, '', totf, gold)
for i, w in enumerate(cw, 1): ws4.column_dimensions[get_column_letter(i)].width = w

# Sheet 5: Credit Suppliers
ws5 = wb.create_sheet('Credit Suppliers')
ws5.merge_cells('A1:F1'); ws5['A1'].value = 'Credit Suppliers — Periodic Statement POs'; ws5['A1'].font = tf; ws5['A1'].alignment = Alignment(horizontal='center')
row = 3; h5 = ['PO #', 'Vendor', 'Amount (SAR)', 'Date', 'Status', 'Note']
for i, h in enumerate(h5, 1): ws5.cell(row=row, column=i, value=h)
hdr(ws5, row, 6)
for idx, r in enumerate(credit_pos):
    row = 4 + idx; alt = lg if idx % 2 == 0 else wf
    sc(ws5, row, 1, r['po'], bfont, alt); sc(ws5, row, 2, r['vendor'], bfont, alt); sc(ws5, row, 3, round(r['amount'],2), bfont, alt, Alignment(horizontal='right'))
    sc(ws5, row, 4, r['date'], bfont, alt, Alignment(horizontal='center')); sc(ws5, row, 5, r['pay_state'], bfont, alt, Alignment(horizontal='center')); sc(ws5, row, 6, 'Credit supplier — balance per periodic statement', bfont, alt)

out = '/Users/mohamedessa/Library/CloudStorage/OneDrive-SAMAYAINVESTMENT/Samaya/Orders/2026/0000 اداريات/00 تقارير الاعمال/Samaya_Factory_Cashout_Report_Updated.xlsx'
wb.save(out)
print(f'\nSaved: {out}', file=sys.stderr)
print(f'\n=== FINAL REPORT ===')
print(f'Cashout Required (unpaid): {len(unpaid)} POs = {total_unpaid:,.2f} SAR')
print(f'Paid via Odoo Bills: {len(bill_paid)} POs = {total_bill_paid:,.2f} SAR')
print(f'Paid Outside Odoo (Ibrahim): {len(chatter_paid)} POs = {total_chatter_paid:,.2f} SAR')
print(f'Credit Suppliers: {len(credit_pos)} POs')
print(f'Grand Total: {grand_total:,.2f} SAR')
