#!/usr/bin/env python3
"""
Design Phase Deliverables Tracker — Overdue / At-Risk / Upcoming report.

Parses the latest "Design Phase Deliverables Tracker*.xlsx" (from CG) across
known locations, computes per-discipline status, and prints a concise report
suitable for a daily cron delivery.

Usage:
    python3 scripts/design_tracker_overdue.py [--xlsx PATH] [--days-at-risk 3] [--days-upcoming 7]

Exit code 0 always (report is informational). Prints the report to stdout.
"""
import argparse
import glob
import os
import sys
from datetime import datetime, date
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent

# Candidate locations for the newest tracker Excel, in priority order.
CANDIDATE_DIRS = [
    Path(os.path.expanduser("~/.hermes/cache/documents")),
    Path(os.path.expanduser("~/.hermes/cache")),
    REPO / "01_Registers",
    Path(os.path.expanduser("~/Desktop")),
    Path(os.path.expanduser("~/Downloads")),
    Path(os.path.expanduser("~/OneDrive - SAMAYA INVESTMENT")),
    Path(os.path.expanduser("~/OneDrive")),
]

# Discipline sheet -> human label. Order matters for the report.
SHEET_LABELS = [
    ("Electrical Deliverables", "Electrical"),
    ("Low Current & ICT Deliverables ", "Low Current & ICT"),
    ("Exhibition Lighting Deliverable", "Exhibition Lighting"),
    ("AV Deliverables", "AV"),
    ("SHOWCASES Deliverables  ", "Showcases"),
    ("BIM Model Deliverables", "BIM"),
    ("scenography Deliverables", "Scenography"),
    ("Arch Deliverables", "Architecture"),
    ("Mech Deliverables ", "Mechanical"),
    ("STR Deliverables ", "Structural"),
]

# Statuses that count as "done" (no longer overdue).
# NOTE: the CG sheet uses "Submitted." (with trailing period) — must be in this set.
DONE_STATUS = {"submitted", "submitted.", "approved", "code-a", "code-b", "code-c", "code-d",
               "final", "closed", "issued", "b", "c", "d", "u", "under review",
               "approved with comments", "approved w/ comments"}

# Statuses that are explicitly "in progress" (still pending).
IN_PROGRESS_STATUS = {"in progress", "inprogress", "in-progress", "in progress "}


def find_latest_xlsx(explicit=None):
    """Return the newest matching xlsx path, or None."""
    if explicit:
        p = Path(explicit)
        return p if p.exists() else None
    best = None
    best_mtime = 0
    for d in CANDIDATE_DIRS:
        if not d.exists():
            continue
        for pat in ("Design Phase Deliverables Tracker*.xlsx",
                    "Design_Phase_Deliverables_Tracker*.xlsx",
                    "*Deliverables_Tracker*.xlsx"):
            for f in glob.glob(str(d / pat)):
                try:
                    m = os.path.getmtime(f)
                except OSError:
                    continue
                if m > best_mtime:
                    best_mtime = m
                    best = f
    return Path(best) if best else None


def _parse_date(v):
    """Return a date or None from a cell value."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y", "%d-%m-%y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _norm(s):
    return (s or "").strip().lower()


def _is_done(status):
    return _norm(status) in DONE_STATUS


def _is_in_progress(status):
    n = _norm(status)
    return n in IN_PROGRESS_STATUS or ("in progress" in n)


def parse_sheet(ws):
    """Return list of dicts: {title, resp, forecast, status, prep, done}."""
    rows = []
    # Find header row (contains 'Forecast' or 'Status' or 'Drawing')
    header_row = None
    header_map = {}
    for r in range(1, min(6, ws.max_row + 1)):
        vals = {}
        for c in range(1, 16):
            v = ws.cell(r, c).value
            if v is not None:
                vals[c] = _norm(str(v))
        joined = " ".join(vals.values())
        if "forecast" in joined or "status" in joined or "drawing" in joined:
            header_row = r
            for c, name in vals.items():
                header_map[name] = c
            break
    if header_row is None:
        return rows

    def col(*names):
        for n in names:
            for k, c in header_map.items():
                if n in k:
                    return c
        return None

    c_title = col("drawing title", "submission description", "drawing package")
    c_resp = col("responsibility")
    c_fore = col("forecast submission", "forecast")
    c_status = col("status")
    c_prep = col("preparation")

    for r in range(header_row + 1, ws.max_row + 1):
        title = ws.cell(r, c_title).value if c_title else None
        if title is None:
            continue
        title = str(title).strip()
        if not title:
            continue
        fore = _parse_date(ws.cell(r, c_fore).value) if c_fore else None
        status = str(ws.cell(r, c_status).value or "").strip() if c_status else ""
        prep = ws.cell(r, c_prep).value if c_prep else None
        resp = str(ws.cell(r, c_resp).value or "").strip() if c_resp else ""
        rows.append({
            "title": title,
            "resp": resp,
            "forecast": fore,
            "status": status,
            "prep": prep,
            "done": _is_done(status),
        })
    return rows


def classify(item, today, at_risk_days, upcoming_days):
    """Return a status bucket for an item."""
    if item["done"]:
        return "done"
    f = item["forecast"]
    if f is None:
        return "no-date"
    delta = (f - today).days
    if delta < 0:
        return "overdue"
    if delta <= at_risk_days:
        return "at-risk"
    if delta <= upcoming_days:
        return "upcoming"
    return "on-track"


def build_report(xlsx, at_risk_days, upcoming_days):
    today = date.today()
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    lines = []
    lines.append(f"# Design Phase Deliverables Tracker — Status ({today.isoformat()})")
    lines.append("")
    lines.append(f"Source: `{xlsx.name}`")
    lines.append("")
    lines.append("Legend: 🔴 Overdue · 🟠 At-risk (≤%dd) · 🟡 Upcoming (≤%dd) · ✅ Done" % (at_risk_days, upcoming_days))
    lines.append("")

    grand = {"overdue": 0, "at-risk": 0, "upcoming": 0, "done": 0, "no-date": 0, "on-track": 0}

    for sheet, label in SHEET_LABELS:
        if sheet not in wb.sheetnames:
            continue
        items = parse_sheet(wb[sheet])
        if not items:
            continue
        buckets = {"overdue": [], "at-risk": [], "upcoming": [], "done": [],
                   "no-date": [], "on-track": []}
        for it in items:
            b = classify(it, today, at_risk_days, upcoming_days)
            buckets[b].append(it)
            grand[b] += 1

        lines.append(f"## {label}  ({len(items)} items)")
        lines.append("")
        lines.append(f"- ✅ Done {len(buckets['done'])} · 🔴 Overdue {len(buckets['overdue'])} · "
                     f"🟠 At-risk {len(buckets['at-risk'])} · 🟡 Upcoming {len(buckets['upcoming'])} · "
                     f"📅 On-track {len(buckets['on-track'])} · ⚪ No date {len(buckets['no-date'])}")
        lines.append("")

        for bucket, icon in (("overdue", "🔴"), ("at-risk", "🟠"), ("upcoming", "🟡")):
            if not buckets[bucket]:
                continue
            lines.append(f"### {icon} {bucket.title()}")
            lines.append("")
            lines.append("| Title | Resp | Forecast | Status | Prep |")
            lines.append("|---|---|---|---|---|")
            for it in sorted(buckets[bucket], key=lambda x: x["forecast"] or date.max):
                prep = f"{it['prep']*100:.0f}%" if isinstance(it["prep"], (int, float)) else (it["prep"] or "—")
                lines.append(f"| {it['title'][:60]} | {it['resp'] or '—'} | "
                             f"{it['forecast'].isoformat() if it['forecast'] else '—'} | "
                             f"{it['status'] or '—'} | {prep} |")
            lines.append("")
        lines.append("")

    lines.append("---")
    lines.append("")
    lines.append("## Overall")
    lines.append("")
    lines.append(f"- 🔴 Overdue: **{grand['overdue']}**")
    lines.append(f"- 🟠 At-risk: **{grand['at-risk']}**")
    lines.append(f"- 🟡 Upcoming: **{grand['upcoming']}**")
    lines.append(f"- ✅ Done: **{grand['done']}**")
    lines.append(f"- 📅 On-track: **{grand['on-track']}**")
    lines.append(f"- ⚪ No date: **{grand['no-date']}**")
    lines.append("")
    return "\n".join(lines)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", help="Explicit path to tracker xlsx")
    ap.add_argument("--days-at-risk", type=int, default=3)
    ap.add_argument("--days-upcoming", type=int, default=7)
    args = ap.parse_args()

    xlsx = find_latest_xlsx(args.xlsx)
    if xlsx is None:
        print("No Design Phase Deliverables Tracker xlsx found in known locations.")
        sys.exit(0)
    print(build_report(xlsx, args.days_at_risk, args.days_upcoming))


if __name__ == "__main__":
    main()
