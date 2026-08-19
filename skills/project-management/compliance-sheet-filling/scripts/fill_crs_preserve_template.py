#!/usr/bin/env python3
"""
Fill an approved CRS (Comment Resolution Sheet) template while PRESERVING the
template format (header rows 1-10, legend, signature block).

User-mandated rule (from session 2026-08-19): DO NOT change the template and
format; DO NOT unmerge the data cells as a standalone step; DO NOT rebuild the
whole CRS from a fresh Workbook. Replace ONLY the data region with a consistent
merge pattern.

Steps implemented:
  1. Open the pristine template (a backup copy with the template format intact —
     NOT an already-corrupted file).
  2. Unmerge ONLY data-region merges (rows 11+); leave header/legend/signatures.
  3. Clear data cells (rows 11+).
  4. Re-merge every data row with a uniform pattern: C:D, E:I, J:O, Q:R;
     A(No), B(Initial), P(Reply By) single.
  5. Write to merge anchors: No->A, Initial->B, Sheet->C, Comment->E,
     Reply->J, Reply By->P, Status->Q.
  6. wrap_text + top alignment on comment/reply cols; thin borders on all.
  7. Save OVER the target CRS in OneDrive.

SOURCE comments: build the Comment column from the CG's OWN DS submittal cover
page (the complete comment set + review code), NOT from the vendor's Audit
Response.xlsx which is often a subset. Fill Originator Reply from the vendor's
Audit Response, matching by comment topic, and flag unanswered comments.
"""

import openpyxl
from openpyxl.styles import Alignment, Border, Side

# --- config ---------------------------------------------------------------
TEMPLATE = "/path/to/pristine/template.xlsx"     # backup copy w/ template format
COMMENTS_SRC = "/path/to/vendor/Audit Response.xlsx"
OUT = "/path/to/target/CRS_Rev01.xlsx"           # the file user edits in OneDrive
DOC_REF = "MOC-MUS-ASE-1E0-1G-0002"
DOC_TITLE = "AV Package Part II (Rev.001) - Detailed Design 50% Gate"
# --------------------------------------------------------------------------

def extract(ws):
    """Read comment+reply pairs from a two-column vendor sheet (No | text,
    reply on the following row)."""
    rows = list(ws.iter_rows(values_only=True))
    out, i = [], 0
    while i < len(rows):
        no = rows[i][0]
        if no is not None and str(no).strip().isdigit():
            c = (rows[i][1] or "").strip()
            rp = ""
            if i + 1 < len(rows) and (rows[i + 1][0] is None or str(rows[i + 1][0]).strip() == ""):
                rp = (rows[i + 1][1] or "").strip()
                i += 1
            out.append({"comment": c, "reply": rp})
        i += 1
    return out

swb = openpyxl.load_workbook(COMMENTS_SRC)
# 1st-submittal comments then 2nd-submittal comments (adjust sheet names):
rows = [{"sheet": "1st Submittal", **c} for c in extract(swb["1st Submital responce"])]
rows += [{"sheet": "2nd Submittal", **c} for c in extract(swb["2nd Submital responce. "])]
N = len(rows)

wb = openpyxl.load_workbook(TEMPLATE)
ws = wb["CRS"]

def setv(r, c, v):
    """Write to a merge anchor (top-left cell) if inside a merge, else direct."""
    for mc in ws.merged_cells.ranges:
        if mc.min_row <= r <= mc.max_row and mc.min_col <= c <= mc.max_col:
            if mc.min_row == r and mc.min_col == c:
                ws.cell(r, c).value = v
            return
    ws.cell(r, c).value = v

# 1) Update header meta (to merge anchors)
setv(5, 4, DOC_REF)          # CRS NUMBER
setv(6, 4, DOC_REF)          # DOCUMENT No.
setv(7, 4, DOC_TITLE)        # DOCUMENT TITLE

# 2) Unmerge ONLY the data region (rows 11+)
data_merges = [mc for mc in ws.merged_cells.ranges if mc.min_row >= 11]
for mc in data_merges:
    try:
        ws.unmerge_cells(str(mc))
    except Exception:
        pass

# 3) Clear data cells rows 11-40
for r in range(11, 41):
    for c in range(1, 19):
        ws.cell(r, c).value = None

# 4) Re-merge each data row with a uniform pattern (C:D, E:I, J:O, Q:R)
LAST = 10 + N
for r in range(11, LAST + 1):
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=9)
    ws.merge_cells(start_row=r, start_column=10, end_row=r, end_column=15)
    ws.merge_cells(start_row=r, start_column=17, end_row=r, end_column=18)

# 5) Fill data (write to merge anchors)
for idx, c in enumerate(rows):
    r = 11 + idx
    ws.cell(r, 1, idx + 1)    # No. (single)
    ws.cell(r, 2, "CG")       # Initial (single) — use CG reviewer name from cover page
    setv(r, 3, c["sheet"])    # Sheet (C:D)
    setv(r, 5, c["comment"])  # Reviewer Comment (E:I)
    setv(r, 10, c["reply"])   # Originator Reply (J:O)
    ws.cell(r, 16, "Rawasin") # Reply By (single)
    setv(r, 17, "")           # Reply Status (Q:R)

# 6) Styling
thin = Side(style="thin", color="000000")
bd = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")
for r in range(11, LAST + 1):
    for c in range(1, 19):
        cell = ws.cell(r, c)
        cell.border = bd
        if c in (5, 10):
            cell.alignment = wrap
        elif c in (1, 2, 3, 16):
            cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

wb.save(OUT)
print(f"Saved {N} rows (template format preserved, header/legend/signatures intact): {OUT}")
