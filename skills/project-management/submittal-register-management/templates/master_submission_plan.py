#!/usr/bin/env python3
"""Template: Master Submission Plan generator with dependency logic and 7-day review buffers.
Copy and adapt for each project's dates and packages."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

wb = openpyxl.Workbook()
hdr_font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
hdr_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
p50_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
par_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
wrap = Alignment(wrap_text=True, vertical='top')
ctr = Alignment(wrap_text=True, vertical='center', horizontal='center')

def ad(d, n):
    return (datetime.strptime(d, '%d/%m/%Y') + timedelta(days=n)).strftime('%d/%m/%Y')

REVIEW = 7

# Define TRACKS: (name, [d50, d90, d100, difc], dependency, parallel_group, notes)
TRACKS = [
    ('Architecture (NRS)', ['29/06/2026', ad('29/06/2026', 30+REVIEW),
     ad('29/06/2026', 60+2*REVIEW), '28/08/2026'], '', 0, 'CRITICAL PATH'),
    ('Acoustic', ['29/06/2026', ad('29/06/2026', 30+REVIEW),
     ad('29/06/2026', 60+2*REVIEW), '28/08/2026'], '', 0, ''),
    ('MEP — Design Basis', ['01/07/2026', '31/07/2026', '30/08/2026', '—'], '', 0, '1st sub'),
    ('MEP — Layouts', ['27/07/2026', '26/08/2026', '25/09/2026', '—'],
     'Needs structural assessment', 1, ''),
    ('MEP — IFC Docs', ['—', '—', '—', '28/08/2026'], '', 5, ''),
    ('AV — Philosophy', ['01/07/2026', '31/07/2026', '30/08/2026', '—'], '', 0, '1st sub'),
    ('AV — System Design', ['29/07/2026', '28/08/2026', '27/09/2026', '—'],
     'Needs Arch 50%', 2, ''),
    ('Lighting — Philosophy', ['10/07/2026', '09/08/2026', '08/09/2026', '—'], '', 0, ''),
    ('Lighting — Layouts', ['10/07/2026', '09/08/2026', '08/09/2026', '—'],
     'Needs Arch 50% ceilings', 2, ''),
    ('Lighting — Conservation', ['—', '—', '—', '—'],
     'BLOCKED — client object list', 4, 'RFI sent Jun 2026'),
    ('Graphics', ['TBC', 'TBC', 'TBC', 'TBC'], 'Client content research', 4, ''),
    ('Model Maker', ['TBC', 'TBC', 'TBC', 'TBC'], 'Client object list', 4, ''),
]

# === SORT by 50% date ===
def sk(t):
    d = t[1][0]
    if d == 'TBC': return '9999-99-99'
    try: return datetime.strptime(d, '%d/%m/%Y').strftime('%Y-%m-%d')
    except: return '9999-99-99'
sorted_tracks = sorted(TRACKS, key=sk)

# === SHEET 1: MASTER PLAN ===
ws = wb.active
ws.title = 'Master Plan'
ws.merge_cells('A1:I1')
c = ws.cell(row=1, column=1, value='MASTER SUBMISSION PLAN')
c.font = Font(name='Calibri', bold=True, size=14, color='1F4E79')

hdrs = ['Package', '50%', '90%', '100%', 'IFC/AFC', 'Depends On', 'Group', 'Review Buffer', 'Notes']
widths = [28, 14, 14, 14, 14, 28, 8, 10, 35]
for i, h in enumerate(hdrs, 1):
    c = ws.cell(row=3, column=i, value=h)
    c.font = hdr_font; c.fill = hdr_fill; c.alignment = ctr; c.border = thin
    ws.column_dimensions[get_column_letter(i)].width = widths[i-1]

row = 4
prev_date = ''
for name, dates, dep, group, notes in sorted_tracks:
    if dates[0] != prev_date and dates[0] != '—':
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(hdrs))
        c = ws.cell(row=row, column=1, value=f'Submit: {dates[0]}')
        c.font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
        c.fill = p50_fill if prev_date == '' else PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
        c.alignment = Alignment(horizontal='left', vertical='center')
        for ci in range(1, len(hdrs)+1): ws.cell(row=row, column=ci).border = thin
        row += 1
        prev_date = dates[0]
    is_parallel = any(t[3] == group and t[1][0] == dates[0] and t[0] != name for t in sorted_tracks)
    f = Font(name='Calibri', size=9, italic=is_parallel)
    fill = par_fill if is_parallel else PatternFill()
    vs = [name, dates[0], dates[1], dates[2], dates[3], dep, str(group), f'{REVIEW}d' if dep else '—', notes]
    for ci, v in enumerate(vs, 1):
        c = ws.cell(row=row, column=ci, value=v)
        c.font = f; c.alignment = wrap; c.border = thin; c.fill = fill
    row += 1

ws.freeze_panes = 'A4'
tmp = '/tmp/Master_Submission_Plan.xlsx'
wb.save(tmp)
print(f'Master Plan saved to {tmp}')
print(f'{len(TRACKS)} packages with {REVIEW}-day review buffers')
