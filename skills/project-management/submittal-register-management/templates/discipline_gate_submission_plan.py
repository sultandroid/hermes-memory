#!/usr/bin/env python3
"""TEMPLATE: Gate-Based Discipline Submission Plan
Copy this file, replace 'Structural' with your discipline, swap item data.
Follows MEP Stage 04 Submission Plan 13-column flat format.

Usage: python3 discipline_gate_submission_plan.py
Output: /tmp/{Discipline}_Submittal_Register.xlsx
"""

import os
from datetime import date, timedelta, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIG ────────────────────────────────────────────────────────────
DISCIPLINE = 'Structural'       # ← change this
OUT_FILE   = f'/tmp/{DISCIPLINE}_Submittal_Register.xlsx'
SHEET_NAME = f'{DISCIPLINE} Stage 04 Plan'
BASE       = date(2026, 6, 29)  # project baseline

# ── Styles (matches MEP template) ─────────────────────────────────────
HDR_FILL  = PatternFill('solid', fgColor='FF366092')
GATE_FILL = PatternFill('solid', fgColor='FF1F4E79')
SYS_FILL  = PatternFill('solid', fgColor='FFBDD7EE')
AMBER     = PatternFill('solid', fgColor='FFF4B942')
HDR_FONT  = Font(name='Calibri', size=10, bold=True, color='FFFFFFFF')
GATE_FONT = Font(name='Calibri', size=12, bold=True, color='FFFFFFFF')
SYS_FONT  = Font(name='Calibri', size=11, bold=True, color='FF000000')
DATA_FONT = Font(name='Calibri', size=10, color='FF000000')
CENTER    = Alignment(horizontal='center', vertical='center', wrap_text=True)
THIN_B    = Border(left=Side('thin', 'FFD9D9D9'), right=Side('thin', 'FFD9D9D9'),
                   top=Side('thin', 'FFD9D9D9'), bottom=Side('thin', 'FFD9D9D9'))

HEADERS = [
    'Gate', 'Level / Zone', 'Discipline', 'Submission Category',
    'Drawing Package / Item', 'Submission Description',
    'Responsibility', 'Planned Submission Date', 'Review Duration (Days)',
    'Approval Authority', 'Linked Activity ID (Program)', 'Status', 'Remarks'
]
COL_W = [18, 24, 14, 22, 42, 50, 16, 22, 20, 22, 28, 16, 40]

# ── Helpers ───────────────────────────────────────────────────────────
def d(y, m, day): return date(y, m, day)
def add7(dt, n=1): return dt + timedelta(days=7 * n)
def add30(dt, n=1): return dt + timedelta(days=30 * n)

def fmt(dt):
    if isinstance(dt, date): return dt.strftime('%d/%m/%Y')
    return str(dt)

def add(gate, zone, cat, pkg, desc, resp, dt_val, review='7',
        approver='CG', link='--', status='Planned', rm=''):
    rows.append([gate, zone, DISCIPLINE, cat, pkg, desc, resp,
                 fmt(dt_val), f'{review} days', approver, link, status, rm])

def add_str(gate, zone, cat, pkg, desc, resp, dt_str, review='7',
            approver='CG', link='--', status='Planned', rm=''):
    rows.append([gate, zone, DISCIPLINE, cat, pkg, desc, resp,
                 dt_str, f'{review} days', approver, link, status, rm])

# ═══════════════════════════════════════════════════════════════════════
# ITEM DATA — Replace with your discipline's items
# ═══════════════════════════════════════════════════════════════════════
rows = []

# ==== Gate 1 — DETAILED DESIGN ====
# Sub-section example: Completed Studies
add('Detailed Design', 'All Levels', 'Detailed Design',
    'ASE-STR-REP-001', 'Sample completed report',
    'Samaya TO', d(2025,3,6), status='Submitted',
    rm='Already submitted, actual date shown.')

# Sub-section example: Pending Items
add('Detailed Design', 'All Levels', 'Detailed Design',
    'ASE-STR-CAL-001', 'Sample pending analysis',
    'Samaya TO', d(2026,7,7), status='In Progress',
    rm='Planned per scope.')

# Gallery-specific per floor
add('Detailed Design', 'Basement (BF)', 'Detailed Design',
    'ASE-STR-GAL-001', 'Per-floor gallery structural item',
    'Samaya TO', d(2026,7,6), rm='Needs Arch GA.')

# ==== Gate 2 — MATERIAL APPROVAL ====
add('Material Approval', 'All Levels', 'Material & Samples Submittals',
    'Material Submittals Register (Live-Sync)',
    'Material Submittals Register',
    'Samaya TO', d(2026,8,25), status='Planned',
    rm='Live-sync register.')

# ==== Gate 3 — COORDINATED IFC ====
for floor, zone_name in [('BF', 'Basement Floor'), ('LGF', 'Lower Ground Floor'),
                         ('GF', 'Ground Floor'), ('1F', 'First Floor')]:
    add('Coordinated IFC', zone_name, 'Coordinated IFC',
        f'STR-300001-IFC-{floor}',
        f'All {DISCIPLINE} disciplines — {zone_name} IFC Package',
        'Samaya TO', d(2026,9,14), review='14', status='Planned')

# ═══════════════════════════════════════════════════════════════════════
# WRITE WORKBOOK
# ═══════════════════════════════════════════════════════════════════════
wb = Workbook()
ws = wb.active
ws.title = SHEET_NAME

# Row 1 — headers
for c, h in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = HDR_FONT; cell.fill = HDR_FILL
    cell.alignment = CENTER; cell.border = THIN_B
ws.row_dimensions[1].height = 30

for i, w in enumerate(COL_W, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Section detection map (key=first item's pkg code marker → section name)
SECTION_MAP = [
    ('REP-', 'Completed Studies & Existing Data'),
    ('CAL-', 'Analysis & Design (per scope)'),
    ('GAL-', 'Gallery-Specific Items'),
    ('RIG-', 'Rigging Systems'),
    ('BIM-', 'BIM Models'),
    ('IFC-', 'All Floors — IFC Packages'),
]

GATE_NAMES = {
    'Detailed Design':   ('Gate 1  —  DETAILED DESIGN  (Assessment & Design)', True),
    'Material Approval': ('Gate 2  —  MATERIAL APPROVAL  (Material Submittals Register)', True),
    'Coordinated IFC':   ('Gate 3  —  COORDINATED IFC  (Issued For Construction)', True),
}

r = 2
gate_headers_written = set()

for row_data in rows:
    gate_val = row_data[0]

    # Insert gate header at first occurrence of each gate
    if gate_val not in gate_headers_written:
        gate_headers_written.add(gate_val)
        gt, has_sub = GATE_NAMES.get(gate_val, (gate_val, False))
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=13)
        cell = ws.cell(row=r, column=1, value=gt)
        cell.font = GATE_FONT; cell.fill = GATE_FILL
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = THIN_B
        for c in range(2, 14):
            ws.cell(row=r, column=c).fill = GATE_FILL
            ws.cell(row=r, column=c).border = THIN_B
        ws.row_dimensions[r].height = 28
        r += 1

    # Insert sub-section header when item's pkg code matches a section marker
    pkg = str(row_data[4])
    for marker, sec_name in SECTION_MAP:
        if marker in pkg:
            # Check if previous row is already this section header
            prev_val = ws.cell(row=r-1, column=1).value if r > 2 else ''
            if prev_val != sec_name:
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=13)
                cell = ws.cell(row=r, column=1, value=sec_name)
                cell.font = SYS_FONT; cell.fill = SYS_FILL
                cell.alignment = Alignment(horizontal='left', vertical='center')
                for c in range(1, 14):
                    ws.cell(row=r, column=c).fill = SYS_FILL
                    ws.cell(row=r, column=c).border = THIN_B
                ws.row_dimensions[r].height = 22
                r += 1
            break

    # Write data row
    rm_val = str(row_data[12])
    for c, v in enumerate(row_data, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = DATA_FONT; cell.alignment = CENTER; cell.border = THIN_B
        if 'Submitted' in rm_val or 'Submitted' in str(row_data[5]):
            cell.fill = AMBER
    ws.row_dimensions[r].height = 20
    r += 1

ws.freeze_panes = 'A2'
wb.save(OUT_FILE)
print(f'Saved: {OUT_FILE} ({len(rows)} items)')
