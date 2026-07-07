# Standard Generator Template

Copy this as a starting point for any new submittal register.

```python
#!/usr/bin/env python3
"""Generate [Discipline] Submittal Register."""
import openpyxl, os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
hf = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
hfl = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
cf = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
sf = Font(name='Calibri', bold=True, size=10, color='375623')
bf = Font(name='Calibri', size=10)
tb = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
wr = Alignment(wrap_text=True, vertical='top')
cw = Alignment(wrap_text=True, vertical='top', horizontal='center')
p50 = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
p90 = PatternFill(start_color='548235', end_color='548235', fill_type='solid')
p100 = PatternFill(start_color='BF8F00', end_color='BF8F00', fill_type='solid')
pifc = PatternFill(start_color='843C0C', end_color='843C0C', fill_type='solid')

hdrs = ['Ref #', 'Submittal / Deliverable', 'Discipline', '50%', '90%', '100%', 'IFC/AFC', 'Sub-Package', 'Remarks']
cww = [7, 50, 14, 14, 14, 14, 14, 22, 28]
def wh(ws):
    for i, h in enumerate(hdrs, 1):
        c = ws.cell(row=1, column=i, value=h); c.font=hf; c.fill=hfl; c.alignment=cw; c.border=tb
        ws.column_dimensions[get_column_letter(i)].width = cww[i-1]

# Items: (ref, desc, disc, [d50, d90, d100, difc], subpkg, rm)
its = [
    ('XX-001', 'Deliverable description', 'Discipline', ['29/06/2026','29/07/2026','28/08/2026','—'], 'Sub-Package', ''),
]

base = '/Users/mohamedessa/Library/CloudStorage/OneDrive-SAMAYAINVESTMENT/Samaya/Technical Office/Bim Unit/Aseer-Museum'
stages = [('50% Design', p50, 0), ('90% Design', p90, 1), ('100% Design', p100, 2), ('IFC  AFC  Construction', pifc, 3)]

# Category name dict: {start_ref_num: 'Category Name', ...}
cn = {1:'A — CATEGORY NAME', N:'B — NEXT CATEGORY'}
cn_keys = sorted(cn.keys())
next_cat = {k: cn_keys[i+1] if i+1 < len(cn_keys) else 999 for i, k in enumerate(cn_keys)}

first = True
for pn, pc, si in stages:
    if first: ws = wb.active; ws.title = pn; first = False
    else: ws = wb.create_sheet(pn)
    wh(ws); row = 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(hdrs))
    c = ws.cell(row=row, column=1, value=f'\u25b2 {pn}')
    c.font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
    c.fill = pc; c.alignment = Alignment(horizontal='left', vertical='center')
    for ci in range(1, len(hdrs)+1):
        ws.cell(row=row, column=ci).border = tb; ws.cell(row=row, column=ci).fill = pc
    row += 1; cnt = 0
    for ref_code, desc, disc, dates, spkg, rm in its:
        rn = int(ref_code.split('-')[1])
        if rn in cn:
            grp = [it for it in its
                   if int(it[0].split('-')[1]) >= rn
                   and int(it[0].split('-')[1]) < next_cat[rn]
                   and it[3][si] != '—']
            if grp:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(hdrs))
                c = ws.cell(row=row, column=1, value=cn[rn])
                c.font = sf; c.fill = cf; c.alignment = Alignment(horizontal='left', vertical='center')
                for ci_ in range(1, len(hdrs)+1):
                    ws.cell(row=row, column=ci_).border = tb; ws.cell(row=row, column=ci_).fill = cf
                row += 1
        if dates[si] == '—': continue
        cnt += 1
        show_dates = ['—','—','—','—']; show_dates[si] = dates[si]
        vs = [ref_code, desc, disc, show_dates[0], show_dates[1], show_dates[2], show_dates[3], spkg, rm]
        for ci_, val in enumerate(vs, 1):
            c = ws.cell(row=row, column=ci_, value=val); c.font = bf; c.alignment = wr; c.border = tb
        row += 1
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(hdrs))
    ws.cell(row=row, column=1, value=f'{cnt} submittal(s)').font = Font(
        name='Calibri', bold=True, italic=True, size=10, color='375623')
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(hdrs))}{row-1}'

wl = wb.create_sheet('Legend')
for ri, r in enumerate([('Column','Description'),('Ref #','XX-xxx'),('Schedule','Description of schedule logic.')]):
    for ci, val in enumerate(r, 1):
        c = wl.cell(row=ri+1, column=ci, value=val)
        c.font = hf if ri == 0 else bf; c.fill = hfl if ri == 0 else PatternFill()
        c.alignment = wr; c.border = tb
wl.column_dimensions['A'].width = 24; wl.column_dimensions['B'].width = 80

# Save
tmp = '/tmp/[Discipline]_Submittal_Register.xlsx'
wb.save(tmp)
print(f'Saved: {tmp}')
for pn, _, si in stages:
    c = sum(1 for it in its if it[3][si] != '—')
    print(f'  {pn}: {c} items')
```
