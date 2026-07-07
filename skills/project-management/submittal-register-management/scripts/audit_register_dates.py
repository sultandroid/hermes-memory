#!/usr/bin/env python3
"""
audit_register_dates.py — Validate date logic in BIM submittal registers.

Checks that O&M/handover/close-out items only have dates in the IFC/AFC column
(never in 50%, 90%, or 100% columns).

Usage:
    python3 audit_register_dates.py <path_to_register.xlsx>
    python3 audit_register_dates.py <directory>  # scans all .xlsx files

Exit code: 0 = all clean, 1 = violations found
"""
import os
import sys
import openpyxl

# Keywords that identify handover/close-out items
HANDOVER_KEYWORDS = [
    'O&M', 'OPERATION', 'MAINTENANCE', 'HANDOVER', 'CLOSE-OUT', 'CLOSEOUT',
    'AS-BUILT', 'AS BUILT', 'RECORD DRAWING', 'RECORD DRAWINGS',
    'TRAINING', 'SPARE', 'SPARES', 'WARRANTY',
    'COMMISSIONING', 'ITP', 'INSPECTION & TEST PLAN',
    'AFC', 'AFC DOCUMENTATION', 'DESIGNER CERTIFICATION',
]

# Stage column indices (0-indexed in the 9-column template)
# Col 4 = 50%, Col 5 = 90%, Col 6 = 100%, Col 7 = IFC/AFC
STAGE_COLS = {'50%': 3, '90%': 4, '100%': 5, 'IFC/AFC': 6}
EARLY_STAGES = ['50%', '90%', '100%']


def audit_register(filepath):
    """Audit a single register file. Returns (filename, violations: list)."""
    violations = []
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        return (os.path.basename(filepath), [f"ERROR: cannot open — {e}"])

    for sn in wb.sheetnames:
        if sn.lower() in ('legend', 'cover', 'summary', 'dashboard'):
            continue
        ws = wb[sn]
        if ws.max_row < 3:
            continue

        # Determine which column is which by reading header row
        headers = {}
        for c in range(1, ws.max_column + 1):
            h = str(ws.cell(row=1, column=c).value or '').strip()
            if h in ('50%', '50% Design'):
                headers['50%'] = c
            elif h in ('90%', '90% Design'):
                headers['90%'] = c
            elif h in ('100%', '100% Design'):
                headers['100%'] = c
            elif h in ('IFC/AFC', 'IFC  AFC  Construction', 'IFC'):
                headers['IFC/AFC'] = c

        if not headers:
            continue  # not a standard register sheet

        for r in range(3, ws.max_row + 1):
            desc = str(ws.cell(row=r, column=2).value or '')
            ref = str(ws.cell(row=r, column=1).value or '')

            # Skip section headers and empty rows
            if not ref or any(ref.startswith(p) for p in ['A —', 'B —', 'C —', 'D —', 'E —', 'F —', 'G —', 'A -', 'B -', 'C -', 'D -', 'E -', 'F -', 'G -']):
                continue

            # Check if this is a handover/close-out item
            is_handover = any(k in desc.upper() for k in HANDOVER_KEYWORDS)
            if not is_handover:
                continue

            # Check early-stage columns
            for stage_name in EARLY_STAGES:
                col = headers.get(stage_name)
                if col is None:
                    continue
                val = ws.cell(row=r, column=col).value
                if val is not None and str(val).strip() not in ('', '—', 'None', 'N/A'):
                    violations.append(
                        f"  [{sn}] R{r} {ref:<20s} | {desc[:50]:<50s} | "
                        f"{stage_name} column has date: {str(val):<20s}"
                    )

    return (os.path.basename(filepath), violations)


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 audit_register_dates.py <file.xlsx> [<file2.xlsx> ...]")
        print("       python3 audit_register_dates.py <directory>")
        sys.exit(1)

    targets = []
    for arg in sys.argv[1:]:
        if os.path.isdir(arg):
            for f in sorted(os.listdir(arg)):
                if f.endswith('.xlsx') and not f.startswith('~'):
                    targets.append(os.path.join(arg, f))
        elif os.path.isfile(arg):
            targets.append(arg)

    if not targets:
        print("No .xlsx files found.")
        sys.exit(1)

    total_violations = 0
    for fp in targets:
        name, violations = audit_register(fp)
        if violations:
            print(f"\n❌ {name} — {len(violations)} violation(s):")
            for v in violations:
                print(v)
            total_violations += len(violations)
        else:
            print(f"✅ {name} — date logic clean")

    if total_violations:
        print(f"\n{total_violations} total violation(s) found.")
        sys.exit(1)
    else:
        print("\nAll registers pass date logic validation.")
        sys.exit(0)


if __name__ == '__main__':
    main()
